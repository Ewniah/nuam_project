"""
Unified View Module for NUAM Calificaciones System
Consolidated from: views.py, views_admin.py, views_factores.py
Migration Date: 2025-11-30
Task: 1.3 - View Unification
Total Functions: 30 (22 routed + 6 utilities + 2 helpers)
Total Routes: 22 (100% preserved)
"""

# ============================================================================
# IMPORTS - PEP 8 Organized
# ============================================================================

# Standard Library (6 imports)
import csv
import io
import json
import logging
from datetime import datetime, timedelta
from decimal import Decimal

# Django Core (12 imports)
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# Third-Party (1 import)
import openpyxl

# Local Application (4 imports)
from .forms import (
    CalificacionTributariaForm,
    InstrumentoFinancieroForm,
    CargaMasivaForm,
    RegistroForm,
    CalificacionFactoresSimpleForm,
)
from .models import (
    CalificacionTributaria,
    InstrumentoFinanciero,
    LogAuditoria,
    CargaMasiva,
    Rol,
    PerfilUsuario,
    IntentoLogin,
    CuentaBloqueada,
    ArchivoCargado,
)
from .permissions import requiere_permiso

# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# Authentication & Security
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION_MINUTES = 30
FAILED_ATTEMPT_WINDOW_MINUTES = 15

# Pagination & Limits
MAX_AUDIT_LOG_RECORDS = 1000
MAX_LOGIN_HISTORY_RECORDS = 50
RECENT_ACTIVITY_DAYS = 7


# ============================================================================
# SECTION 1: UTILITIES AND HELPERS
# ============================================================================
# Functions: obtener_ip_cliente, verificar_cuenta_bloqueada,
#            registrar_intento_login, verificar_intentos_fallidos,
#            procesar_excel, procesar_csv
# Lines: 52-250 (approx. 200 lines)
# ============================================================================


def obtener_ip_cliente(request):
    """
    Obtiene la dirección IP real del cliente considerando proxies y balanceadores de carga.

    Esta función prioriza el header X-Forwarded-For para obtener la IP original del cliente
    cuando la aplicación está detrás de un proxy o load balancer. Si no existe, utiliza
    REMOTE_ADDR como fallback.

    Parámetros:
        request (HttpRequest): El objeto de solicitud HTTP de Django.

    Retorna:
        str: La dirección IP del cliente (ej: "192.168.1.1").

    Notas:
        - X-Forwarded-For puede contener múltiples IPs separadas por comas; se toma la primera.
        - REMOTE_ADDR contiene la IP del último hop (proxy si existe).
        - Utilizado para auditoría en LogAuditoria e IntentoLogin.
    """
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0]
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


def verificar_cuenta_bloqueada(username):
    """
    Verifica si una cuenta de usuario está bloqueada por intentos fallidos de login.

    Comprueba la tabla CuentaBloqueada para determinar si el usuario está bloqueado
    y calcula los minutos restantes hasta el desbloqueo automático.

    Parámetros:
        username (str): El nombre de usuario a verificar.

    Retorna:
        tuple: Una tupla (bloqueada, mensaje, minutos_restantes) donde:
            - bloqueada (bool): True si la cuenta está actualmente bloqueada.
            - mensaje (str): Mensaje descriptivo del estado del bloqueo.
            - minutos_restantes (int): Minutos hasta el desbloqueo automático.

    Notas:
        - El tiempo de bloqueo está definido por LOCKOUT_DURATION_MINUTES (30 min por defecto).
        - Si el tiempo de bloqueo ha expirado, se desbloquea automáticamente.
        - Retorna (False, "", 0) si la cuenta no existe o no está bloqueada.
    """
    try:
        user = User.objects.get(username=username)
        cuenta_bloqueada = CuentaBloqueada.objects.filter(usuario=user, bloqueada=True).first()

        if cuenta_bloqueada:
            # Verificar si ya pasó el tiempo de bloqueo
            tiempo_bloqueo = timezone.now() - cuenta_bloqueada.fecha_bloqueo
            MINUTOS_BLOQUEO = LOCKOUT_DURATION_MINUTES

            if tiempo_bloqueo.total_seconds() < MINUTOS_BLOQUEO * 60:
                minutos_restantes = int(
                    (MINUTOS_BLOQUEO * 60 - tiempo_bloqueo.total_seconds()) / 60
                )
                return (
                    True,
                    f"Cuenta bloqueada. Intente nuevamente en {minutos_restantes} minutos.",
                    minutos_restantes,
                )
            else:
                # ✅ ACTUALIZADO: Desbloquear automáticamente y registrar en auditoría
                cuenta_bloqueada.bloqueada = False
                cuenta_bloqueada.fecha_desbloqueo = timezone.now()
                cuenta_bloqueada.save()

                # Registrar desbloqueo automático en auditoría
                LogAuditoria.objects.create(
                    usuario=user,
                    accion="ACCOUNT_UNLOCKED",
                    tabla_afectada="CuentaBloqueada",
                    registro_id=cuenta_bloqueada.id,
                    ip_address="SYSTEM",  # Sistema automático
                    detalles=f"Cuenta de {user.username} desbloqueada automáticamente después de 30 minutos",
                )

                return False, "", 0

        return False, "", 0
    except User.DoesNotExist:
        return False, "", 0


def registrar_intento_login(username, ip_address, exitoso, detalles=""):
    """
    Registra todos los intentos de autenticación en la tabla IntentoLogin para auditoría.

    Función auxiliar utilizada por login_view para mantener histórico completo de intentos
    de login exitosos y fallidos, facilitando análisis de seguridad y debugging.

    Parámetros:
        username (str): Nombre de usuario que intentó autenticarse.
        ip_address (str): Dirección IP desde donde se realizó el intento.
        exitoso (bool): True si la autenticación fue exitosa, False si falló.
        detalles (str, opcional): Información adicional sobre el intento. Por defecto: "".

    Retorna:
        None: La función no retorna valor, solo crea registro en BD.

    Notas:
        - Llamada automáticamente por login_view en cada intento
        - Utilizada para análisis de patrones de ataque
        - Base de datos para verificar_intentos_fallidos
        - Visible en admin panel de gestión de usuarios
    """
    IntentoLogin.objects.create(
        username=username, ip_address=ip_address, exitoso=exitoso, detalles=detalles
    )


def verificar_intentos_fallidos(username, ip_address):
    """
    Verifica intentos fallidos recientes y bloquea la cuenta si excede el umbral.

    Cuenta los intentos fallidos de login en la ventana de tiempo definida por
    FAILED_ATTEMPT_WINDOW_MINUTES. Si alcanza MAX_LOGIN_ATTEMPTS, bloquea automáticamente
    la cuenta y registra la acción en auditoría.

    Parámetros:
        username (str): Nombre de usuario a verificar.
        ip_address (str): IP desde donde se realizó el último intento.

    Retorna:
        tuple: (debe_bloquear, intentos) donde:
            - debe_bloquear (bool): True si se alcanzó el límite y se bloqueó la cuenta.
            - intentos (int): Número de intentos fallidos en la ventana de tiempo.

    Notas:
        - Ventana de tiempo: FAILED_ATTEMPT_WINDOW_MINUTES (15 minutos por defecto)
        - Umbral de bloqueo: MAX_LOGIN_ATTEMPTS (5 intentos por defecto)
        - Crea o actualiza registro en CuentaBloqueada
        - Registra acción ACCOUNT_LOCKED en LogAuditoria
        - Retorna (False, intentos) si el usuario no existe
    """
    INTENTOS_MAXIMOS = MAX_LOGIN_ATTEMPTS
    VENTANA_TIEMPO = FAILED_ATTEMPT_WINDOW_MINUTES

    tiempo_limite = timezone.now() - timedelta(minutes=VENTANA_TIEMPO)

    # Contar intentos fallidos recientes
    intentos_fallidos = IntentoLogin.objects.filter(
        username=username, exitoso=False, fecha_hora__gte=tiempo_limite
    ).count()

    if intentos_fallidos >= INTENTOS_MAXIMOS:
        try:
            user = User.objects.get(username=username)

            # Crear o actualizar registro de cuenta bloqueada
            cuenta_bloqueada, created = CuentaBloqueada.objects.get_or_create(
                usuario=user,
                defaults={
                    "intentos_fallidos": intentos_fallidos,
                    "bloqueada": True,
                    "razon": f"Bloqueada automáticamente por {intentos_fallidos} intentos fallidos",
                },
            )

            if not created:
                cuenta_bloqueada.bloqueada = True
                cuenta_bloqueada.intentos_fallidos = intentos_fallidos
                cuenta_bloqueada.fecha_bloqueo = timezone.now()
                cuenta_bloqueada.save()

            # Registrar en auditoría
            LogAuditoria.objects.create(
                usuario=user,
                accion="ACCOUNT_LOCKED",
                tabla_afectada="CuentaBloqueada",
                registro_id=cuenta_bloqueada.id,
                ip_address=ip_address,
                detalles=f"Cuenta bloqueada por {intentos_fallidos} intentos fallidos",
            )

            return True, intentos_fallidos
        except User.DoesNotExist:
            pass

    return False, intentos_fallidos


def procesar_excel(archivo):
    """
    Procesa archivo Excel (.xlsx) y extrae registros para carga masiva.

    Lee archivo Excel, extrae headers de la primera fila y convierte cada fila
    subsecuente en un diccionario. Filtra filas sin código de instrumento.

    Parámetros:
        archivo (UploadedFile): Archivo Excel cargado desde request.FILES.

    Retorna:
        list[dict]: Lista de diccionarios donde cada dict representa una fila con
            headers como keys. Solo incluye filas con codigo_instrumento no vacío.

    Notas:
        - Librería: openpyxl
        - Lee la hoja activa del workbook
        - Primera fila debe contener headers (nombres de columnas)
        - Filas sin codigo_instrumento se omiten automáticamente
        - Utilizada por carga_masiva() para importación batch
    """
    wb = openpyxl.load_workbook(archivo)
    sheet = wb.active
    registros = []

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        registro = dict(zip(headers, row))
        if registro.get("codigo_instrumento"):
            registros.append(registro)

    return registros


def procesar_csv(archivo):
    """
    Procesa archivo CSV y extrae registros para carga masiva.

    Lee archivo CSV codificado en UTF-8 y convierte cada fila en un diccionario
    usando los headers de la primera línea como keys.

    Parámetros:
        archivo (UploadedFile): Archivo CSV cargado desde request.FILES.

    Retorna:
        list[dict]: Lista de diccionarios donde cada dict representa una fila CSV.

    Notas:
        - Librería: csv (stdlib)
        - Encoding: UTF-8 (debe especificarse en archivo)
        - Primera línea debe contener headers (nombres de columnas)
        - Separador: coma (,)
        - Utilizada por carga_masiva() para importación batch
        - No filtra filas vacías (lo hace carga_masiva)
    """
    contenido = archivo.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(contenido))
    return list(reader)


# ============================================================================
# SECTION 2: AUTHENTICATION AND SECURITY
# ============================================================================
# Functions: login_view, logout_view
# Lines: 251-370 (approx. 120 lines)
# ============================================================================


def login_view(request):
    """
    Vista de autenticación de usuarios con auditoría completa y sistema de bloqueo por intentos fallidos.

    Gestiona el proceso de login incluyendo:
    - Verificación de cuenta bloqueada
    - Autenticación de credenciales
    - Registro de intentos exitosos y fallidos en IntentoLogin
    - Registro de auditoría en LogAuditoria
    - Bloqueo automático después de MAX_LOGIN_ATTEMPTS intentos fallidos
    - Desbloqueo automático de cuentas tras login exitoso

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP con datos POST (username, password).

    Retorna:
        HttpResponse:
            - Redirect a 'dashboard' si login exitoso.
            - Render de 'registration/login.html' si cuenta bloqueada o credenciales inválidas.

    Notas:
        - Sistema de bloqueo: 5 intentos fallidos → 30 minutos de bloqueo
        - Ventana de tiempo para contar intentos: 15 minutos
        - Registra acciones: LOGIN (exitoso), LOGIN_FAILED (fallido)
        - Muestra advertencias al usuario cuando quedan ≤2 intentos
    """
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        ip_address = obtener_ip_cliente(request)

        logger.debug(f"Login attempt - Username: {username}, IP: {ip_address}")

        # 1. Verificar si la cuenta está bloqueada
        bloqueada, mensaje_bloqueo, minutos = verificar_cuenta_bloqueada(username)
        if bloqueada:
            logger.warning(
                f"Blocked login attempt - User: {username}, IP: {ip_address}, "
                f"Remaining lockout: {minutos} minutes"
            )
            messages.error(request, mensaje_bloqueo)
            registrar_intento_login(username, ip_address, False, "Intento en cuenta bloqueada")
            return render(request, "registration/login.html")

        # 2. Intentar autenticar
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Login exitoso
            login(request, user)

            logger.info(
                f"Successful login - User: {user.username}, Name: {user.get_full_name()}, "
                f"IP: {ip_address}, Role: {user.groups.first().name if user.groups.exists() else 'No role'}"
            )

            # Registrar intento exitoso
            registrar_intento_login(username, ip_address, True, "Login exitoso")

            # Registrar en auditoría
            LogAuditoria.objects.create(
                usuario=user,
                accion="LOGIN",
                tabla_afectada="User",
                registro_id=user.id,
                ip_address=ip_address,
                detalles=f"Login exitoso desde {ip_address}",
            )

            # Limpiar bloqueo si existía
            CuentaBloqueada.objects.filter(usuario=user).update(
                bloqueada=False, intentos_fallidos=0, fecha_desbloqueo=timezone.now()
            )

            messages.success(request, f"¡Bienvenido {user.first_name or user.username}!")
            return redirect("dashboard")
        else:
            # Login fallido
            logger.warning(
                f"Failed login attempt - Username: {username}, IP: {ip_address}, "
                f"Reason: Invalid credentials"
            )
            registrar_intento_login(username, ip_address, False, "Credenciales incorrectas")

            # Registrar en auditoría (sin usuario porque falló)
            LogAuditoria.objects.create(
                usuario=None,
                accion="LOGIN_FAILED",
                tabla_afectada="User",
                ip_address=ip_address,
                detalles=f"Intento fallido para usuario: {username}",
            )

            # Verificar si debe bloquearse la cuenta
            debe_bloquear, intentos = verificar_intentos_fallidos(username, ip_address)

            if debe_bloquear:
                logger.error(
                    f"Account locked - Username: {username}, IP: {ip_address}, "
                    f"Failed attempts: {intentos}, Lockout duration: {LOCKOUT_DURATION_MINUTES} minutes"
                )
                messages.error(
                    request,
                    f"Cuenta bloqueada por múltiples intentos fallidos. "
                    f"Intente nuevamente en 30 minutos.",
                )
            else:
                intentos_restantes = 5 - intentos
                if intentos_restantes <= 2:
                    messages.warning(
                        request,
                        f"Credenciales incorrectas. Le quedan {intentos_restantes} intentos "
                        f"antes de que su cuenta sea bloqueada.",
                    )
                else:
                    messages.error(request, "Credenciales incorrectas. Intente nuevamente.")

    return render(request, "registration/login.html")


@login_required
def logout_view(request):
    """
    Cierra la sesión del usuario y registra la acción en auditoría.

    Termina la sesión activa del usuario, registra el evento en LogAuditoria y
    redirige a la página de inicio.

    Parámetros:
        request (HttpRequest): Solicitud HTTP con usuario autenticado.

    Retorna:
        HttpResponse: Redirect a 'home' con mensaje informativo.

    Notas:
        - Requiere autenticación: @login_required
        - Registra acción LOGOUT en LogAuditoria con IP del cliente
        - Logging: INFO con username e IP
        - Limpia la sesión de Django completamente
        - Mensaje informativo al usuario tras logout
    """
    user = request.user
    ip_address = obtener_ip_cliente(request)

    logger.info(f"User logout - User: {user.username}, IP: {ip_address}")

    LogAuditoria.objects.create(
        usuario=request.user,
        accion="LOGOUT",
        tabla_afectada="User",
        registro_id=request.user.id,
        ip_address=ip_address,
        detalles=f"Logout desde {ip_address}",
    )

    logout(request)
    messages.info(request, "Sesión cerrada correctamente.")
    return redirect("login")


# ============================================================================
# SECTION 3: DASHBOARD AND REPORTING
# ============================================================================
# Functions: dashboard
# Lines: 371-470 (approx. 100 lines)
# ============================================================================


@login_required
@requiere_permiso("consultar")
def dashboard(request):
    """
    Dashboard principal con diseño de 3 zonas: Acciones Rápidas, KPIs/Charts, Auditoría.

    Presenta un dashboard moderno optimizado para eficiencia del usuario:
    - ZONA A: Acciones rápidas (Nueva Calificación, Carga Masiva, Buscar/Listar)
    - ZONA B: KPIs totales + Charts interactivos (distribución mercado, cargas recientes)
    - ZONA C: Registro de actividad reciente (últimos 10 logs de auditoría)

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP del usuario autenticado.

    Retorna:
        HttpResponse: Render de 'calificaciones/dashboard.html' con context dict conteniendo
            métricas, datos de gráficos y logs de auditoría.

    Notas:
        - Requiere autenticación: @login_required
        - Requiere permiso: @requiere_permiso("consultar")
        - Charts implementados con Chart.js (CDN incluido en template)
        - Datos preparados para Doughnut chart (mercado) y Bar chart (cargas)
    """
    from datetime import datetime
    from django.db.models import Count

    logger.debug(
        f"Dashboard access - User: {request.user.username}, IP: {obtener_ip_cliente(request)}"
    )

    # ========== ZONA A: No requiere datos (solo enlaces estáticos) ==========

    # ========== ZONA B: KPIs y Datos de Charts ==========
    
    # KPI: Totales
    total_calificaciones = CalificacionTributaria.objects.filter(activo=True).count()
    total_instrumentos = InstrumentoFinanciero.objects.filter(activo=True).count()
    total_usuarios = User.objects.filter(is_active=True).count()
    cargas_hoy = CargaMasiva.objects.filter(
        fecha_carga__date=timezone.now().date()
    ).count()
    
    # KPI: Calificaciones del mes actual
    inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    calificaciones_mes = CalificacionTributaria.objects.filter(
        activo=True,
        fecha_creacion__gte=inicio_mes
    ).count()
    
    # KPI: Calificaciones de esta semana
    inicio_semana = timezone.now().date() - timedelta(days=timezone.now().weekday())
    calificaciones_semana = CalificacionTributaria.objects.filter(
        activo=True,
        fecha_creacion__date__gte=inicio_semana
    ).count()

    logger.info(
        f"Dashboard KPIs - User: {request.user.username}, "
        f"Calificaciones: {total_calificaciones}, Instrumentos: {total_instrumentos}, "
        f"Cargas hoy: {cargas_hoy}, Mes: {calificaciones_mes}, Semana: {calificaciones_semana}"
    )

    # Chart 1: Distribución por Mercado (Doughnut)
    mercado_stats = (
        CalificacionTributaria.objects.filter(activo=True)
        .values("mercado")
        .annotate(total=Count("id"))
        .order_by("mercado")
    )
    labels_mercado = [item["mercado"] or "Sin Mercado" for item in mercado_stats]
    data_mercado = [item["total"] for item in mercado_stats]

    # Chart 2: Cargas de los últimos 7 días (Bar)
    labels_cargas = []
    data_cargas = []
    for i in range(6, -1, -1):  # De 6 días atrás a hoy
        fecha = timezone.now().date() - timedelta(days=i)
        labels_cargas.append(fecha.strftime("%d/%m"))
        cargas_dia = CargaMasiva.objects.filter(fecha_carga__date=fecha).count()
        data_cargas.append(cargas_dia)
    
    # Chart 3: Distribución por Origen/Tipo Sociedad (Pie)
    origen_stats = (
        CalificacionTributaria.objects.filter(activo=True)
        .values("tipo_sociedad")
        .annotate(total=Count("id"))
        .order_by("tipo_sociedad")
    )
    labels_origen = [
        "Corredora" if item["tipo_sociedad"] == "A" 
        else "Bolsa" if item["tipo_sociedad"] == "C" 
        else "Sin Tipo" 
        for item in origen_stats
    ]
    data_origen = [item["total"] for item in origen_stats]
    
    # Chart 4: Top 5 Instrumentos Más Usados (Horizontal Bar)
    top_instrumentos = (
        CalificacionTributaria.objects.filter(activo=True)
        .values("instrumento__codigo_instrumento")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )
    labels_top_instrumentos = [item["instrumento__codigo_instrumento"] or "Sin Código" for item in top_instrumentos]
    data_top_instrumentos = [item["total"] for item in top_instrumentos]

    # ========== ZONA C: Actividad Reciente ==========
    
    # Últimos 10 logs de auditoría
    ultimos_logs = LogAuditoria.objects.select_related("usuario").order_by("-fecha_hora")[:10]

    context = {
        # KPIs
        "total_calificaciones": total_calificaciones,
        "total_instrumentos": total_instrumentos,
        "total_usuarios": total_usuarios,
        "cargas_hoy": cargas_hoy,
        "calificaciones_mes": calificaciones_mes,
        "calificaciones_semana": calificaciones_semana,
        
        # Chart 1: Distribución Mercado
        "labels_mercado": labels_mercado,
        "data_mercado": data_mercado,
        
        # Chart 2: Cargas Recientes
        "labels_cargas": labels_cargas,
        "data_cargas": data_cargas,
        
        # Chart 3: Distribución Origen
        "labels_origen": labels_origen,
        "data_origen": data_origen,
        
        # Chart 4: Top Instrumentos
        "labels_top_instrumentos": labels_top_instrumentos,
        "data_top_instrumentos": data_top_instrumentos,
        
        # Auditoría
        "ultimos_logs": ultimos_logs,
        
        # Metadata
        "today": timezone.now(),
    }

    return render(request, "calificaciones/dashboard.html", context)


# ============================================================================
# SECTION 4: CALIFICACIONES CRUD OPERATIONS
# ============================================================================
# Functions: listar_calificaciones, crear_calificacion, editar_calificacion,
#            eliminar_calificacion, crear_calificacion_factores,
#            editar_calificacion_factores
# Lines: 471-650 (approx. 180 lines)
# ============================================================================


@login_required
@requiere_permiso("consultar")
def listar_calificaciones(request):
    """
    Lista todas las calificaciones tributarias activas con filtrado avanzado y paginación.

    Muestra tabla paginada de calificaciones con opciones de filtrado por mercado, origen,
    ejercicio, código de instrumento, rango de fechas y número de DJ. Optimizado para
    hardware con CPU limitado mediante paginación server-side (50 registros por página).

    Parámetros:
        request (HttpRequest): GET request con parámetros opcionales:
            - mercado (str): Filtro por tipo de mercado (ACN, CFI, FFM).
            - tipo_sociedad (str): Filtro por origen (A=Corredora, C=Bolsa).
            - ejercicio (int): Filtro por año de ejercicio comercial.
            - codigo_instrumento (str): Filtro parcial por código (ICONTAINS).
            - fecha_desde (str): Fecha mínima del informe (formato YYYY-MM-DD).
            - fecha_hasta (str): Fecha máxima del informe (formato YYYY-MM-DD).
            - numero_dj (str): Filtro parcial por número de DJ (ICONTAINS).
            - page (int): Número de página para paginación.

    Retorna:
        HttpResponse: Render de 'calificaciones/listar.html' con:
            - calificaciones: QuerySet paginado de CalificacionTributaria
            - page_obj: Objeto Page de Django Paginator
            - Todos los parámetros de filtros en context para mantener estado

    Notas:
        - Solo muestra registros con activo=True (borrado lógico)
        - Ordenado por fecha_creacion descendente (más recientes primero)
        - Paginación: 50 registros por página (optimización para CPU limitado)
        - Query optimizado con select_related('instrumento', 'usuario_creador')
        - Template: 'calificaciones/listar.html' con sticky columns CSS
    """
    # Base queryset con optimización de queries
    calificaciones = CalificacionTributaria.objects.filter(activo=True).select_related(
        "instrumento", "usuario_creador"
    )

    # NUEVOS FILTROS - Metadata fields (Phase 3)
    mercado = request.GET.get("mercado", "").strip()
    tipo_sociedad = request.GET.get("tipo_sociedad", "").strip()
    ejercicio = request.GET.get("ejercicio", "").strip()

    if mercado:
        calificaciones = calificaciones.filter(mercado__iexact=mercado)

    if tipo_sociedad:
        calificaciones = calificaciones.filter(tipo_sociedad__iexact=tipo_sociedad)

    if ejercicio:
        try:
            ejercicio_int = int(ejercicio)
            calificaciones = calificaciones.filter(ejercicio=ejercicio_int)
        except ValueError:
            logger.warning(f"Invalid ejercicio filter value: {ejercicio}")

    # FILTROS EXISTENTES (legacy compatibility)
    codigo_instrumento = request.GET.get("codigo_instrumento", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    numero_dj = request.GET.get("numero_dj", "").strip()

    if codigo_instrumento:
        calificaciones = calificaciones.filter(
            instrumento__codigo_instrumento__icontains=codigo_instrumento
        )

    if fecha_desde:
        calificaciones = calificaciones.filter(fecha_informe__gte=fecha_desde)

    if fecha_hasta:
        calificaciones = calificaciones.filter(fecha_informe__lte=fecha_hasta)

    if numero_dj:
        calificaciones = calificaciones.filter(numero_dj__icontains=numero_dj)

    # Ordenamiento
    calificaciones = calificaciones.order_by("-fecha_creacion")

    # PAGINACIÓN - Server-side (50 records per page for CPU optimization)
    paginator = Paginator(calificaciones, 50)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    logger.info(
        f"Calificaciones list - User: {request.user.username}, "
        f"Total: {paginator.count}, Page: {page_number}/{paginator.num_pages}, "
        f"Filters: mercado={mercado}, tipo_sociedad={tipo_sociedad}, ejercicio={ejercicio}"
    )

    context = {
        "calificaciones": page_obj,  # Paginado
        "page_obj": page_obj,
        # New filters
        "mercado": mercado,
        "tipo_sociedad": tipo_sociedad,
        "ejercicio": ejercicio,
        # Legacy filters
        "codigo_instrumento": codigo_instrumento,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "numero_dj": numero_dj,
    }

    return render(request, "calificaciones/listar.html", context)


@login_required
@requiere_permiso("crear")
def crear_calificacion(request):
    """
    Crea una nueva calificación tributaria con validación y auditoría.

    Presenta formulario para crear calificación manualmente. Valida datos, asigna usuario
    creador automáticamente y registra la operación en LogAuditoria.

    Parámetros:
        request (HttpRequest):
            - GET: Muestra formulario vacío
            - POST: Procesa formulario con datos de calificación

    Retorna:
        HttpResponse:
            - GET: Render de 'calificaciones/form_calificacion.html' con formulario vacío
            - POST exitoso: Redirect a 'listar_calificaciones' con mensaje de éxito
            - POST con error: Render de form con errores de validación

    Notas:
        - Requiere permiso: 'crear'
        - Usuario creador se asigna automáticamente (no editable)
        - Registra acción CREATE en LogAuditoria con IP del cliente
        - Maneja IntegrityError, ValidationError y excepciones genéricas
        - Logging: INFO en creación exitosa, ERROR en fallos
    """
    if request.method == "POST":
        form = CalificacionTributariaForm(request.POST)
        if form.is_valid():
            try:
                calificacion = form.save(commit=False)
                calificacion.usuario_creador = request.user
                calificacion.save()

                logger.info(
                    f"Calificacion created - User: {request.user.username}, "
                    f"ID: {calificacion.id}, Instrumento: {calificacion.instrumento.codigo_instrumento}, "
                    f"Monto: {calificacion.monto}, Factor: {calificacion.factor}"
                )

                # Registrar en auditoría
                ip_address = obtener_ip_cliente(request)
                LogAuditoria.objects.create(
                    usuario=request.user,
                    accion="CREATE",
                    tabla_afectada="CalificacionTributaria",
                    registro_id=calificacion.id,
                    ip_address=ip_address,
                    detalles=f"Calificación creada: {calificacion.instrumento.codigo_instrumento}",
                )

                messages.success(request, "Calificación creada exitosamente.")
                return redirect("listar_calificaciones")
            except IntegrityError as e:
                logger.error(f"Database integrity error creating calificacion: {e}", exc_info=True)
                messages.error(
                    request,
                    "Error de integridad: La calificación ya existe o hay datos duplicados.",
                )
            except ValidationError as e:
                logger.warning(f"Validation error creating calificacion: {e}")
                messages.error(request, f"Error de validación: {e}")
            except Exception as e:
                logger.error(f"Unexpected error creating calificacion: {e}", exc_info=True)
                messages.error(request, f"Error inesperado al crear calificación: {str(e)}")
    else:
        form = CalificacionTributariaForm()

    # Crear mapa de instrumentos para autoselección de DJ
    instrumentos_map = {}
    for instrumento in InstrumentoFinanciero.objects.filter(activo=True):
        # Regla de negocio:
        # - Acción/Sociedad Anónima -> DJ 1949
        # - Fondo (CFI, FM, etc.) -> DJ 1922
        tipo_lower = instrumento.tipo_instrumento.lower()
        if 'accion' in tipo_lower or 'sociedad' in tipo_lower or 'sa' in tipo_lower:
            dj_recomendado = '1949'
        elif 'fondo' in tipo_lower or 'cfi' in tipo_lower or 'fm' in tipo_lower:
            dj_recomendado = '1922'
        else:
            dj_recomendado = '1949'  # Default
        
        instrumentos_map[str(instrumento.id)] = {
            'tipo': instrumento.tipo_instrumento,
            'dj': dj_recomendado
        }

    # Convertir a JSON para pasar al template
    instrumentos_map_json = json.dumps(instrumentos_map)

    return render(request, "calificaciones/form_calificacion.html", {
        "form": form,
        "instrumentos_map_json": instrumentos_map_json
    })


@login_required
@requiere_permiso("modificar")
def editar_calificacion(request, pk):
    """
    Edita una calificación tributaria existente con validación y auditoría.

    Permite modificar todos los campos de una calificación previamente creada. Valida
    que el registro exista y esté activo antes de permitir edición.

    Parámetros:
        request (HttpRequest):
            - GET: Muestra formulario pre-poblado con datos actuales
            - POST: Procesa formulario con datos actualizados
        pk (int): Primary key de la CalificacionTributaria a editar.

    Retorna:
        HttpResponse:
            - GET: Render de 'calificaciones/form_calificacion.html' con form poblado
            - POST exitoso: Redirect a 'listar_calificaciones' con mensaje de éxito
            - POST con error: Render de form con errores, sin guardar cambios

    Excepciones:
        Http404: Si la calificación no existe o está inactiva (activo=False).

    Notas:
        - Requiere permiso: 'modificar'
        - Registra acción UPDATE en LogAuditoria con IP del cliente
        - Maneja IntegrityError, ValidationError y excepciones genéricas
        - Logging: INFO en actualización exitosa, WARNING/ERROR en fallos
        - Usuario creador original no se modifica
    """
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, activo=True)

    if request.method == "POST":
        form = CalificacionTributariaForm(request.POST, instance=calificacion)
        if form.is_valid():
            try:
                calificacion = form.save()

                logger.info(
                    f"Calificacion updated - User: {request.user.username}, "
                    f"ID: {calificacion.id}, Instrumento: {calificacion.instrumento.codigo_instrumento}"
                )

                # Registrar en auditoría
                ip_address = obtener_ip_cliente(request)
                LogAuditoria.objects.create(
                    usuario=request.user,
                    accion="UPDATE",
                    tabla_afectada="CalificacionTributaria",
                    registro_id=calificacion.id,
                    ip_address=ip_address,
                    detalles=f"Calificación editada: {calificacion.instrumento.codigo_instrumento}",
                )

                messages.success(request, "Calificación actualizada exitosamente.")
                return redirect("listar_calificaciones")
            except IntegrityError as e:
                logger.error(
                    f"Database integrity error updating calificacion {pk}: {e}", exc_info=True
                )
                messages.error(request, "Error de integridad al actualizar la calificación.")
            except ValidationError as e:
                logger.warning(f"Validation error updating calificacion {pk}: {e}")
                messages.error(request, f"Error de validación: {e}")
            except Exception as e:
                logger.error(f"Unexpected error updating calificacion {pk}: {e}", exc_info=True)
                messages.error(request, f"Error inesperado: {str(e)}")
    else:
        form = CalificacionTributariaForm(instance=calificacion)

    # Crear mapa de instrumentos para autoselección de DJ (igual que en crear_calificacion)
    instrumentos_map = {}
    for instrumento in InstrumentoFinanciero.objects.filter(activo=True):
        tipo_lower = instrumento.tipo_instrumento.lower()
        if 'accion' in tipo_lower or 'sociedad' in tipo_lower or 'sa' in tipo_lower:
            dj_recomendado = '1949'
        elif 'fondo' in tipo_lower or 'cfi' in tipo_lower or 'fm' in tipo_lower:
            dj_recomendado = '1922'
        else:
            dj_recomendado = '1949'
        
        instrumentos_map[str(instrumento.id)] = {
            'tipo': instrumento.tipo_instrumento,
            'dj': dj_recomendado
        }

    instrumentos_map_json = json.dumps(instrumentos_map)

    context = {
        "form": form, 
        "calificacion": calificacion,
        "instrumentos_map_json": instrumentos_map_json
    }
    return render(request, "calificaciones/form_calificacion.html", context)


@login_required
@requiere_permiso("eliminar")
def eliminar_calificacion(request, pk):
    """
    Eliminación lógica (soft delete) de una calificación tributaria.

    No elimina físicamente el registro de la base de datos, sino que marca el campo
    activo=False, preservando la integridad de auditoría e históricos.

    Parámetros:
        request (HttpRequest):
            - GET: Muestra página de confirmación
            - POST: Ejecuta la eliminación lógica
        pk (int): Primary key de la CalificacionTributaria a eliminar.

    Retorna:
        HttpResponse:
            - GET: Render de 'calificaciones/confirmar_eliminar.html' con objeto
            - POST exitoso: Redirect a 'listar_calificaciones' con mensaje de éxito
            - POST con error: Redirect a 'listar_calificaciones' con mensaje de error

    Excepciones:
        Http404: Si la calificación no existe o ya está inactiva.

    Notas:
        - Requiere permiso: 'eliminar'
        - Eliminación lógica: activo=False (registro permanece en BD)
        - Registra acción DELETE en LogAuditoria con IP del cliente
        - Logging: WARNING en eliminación exitosa, ERROR en fallos
        - No afecta relaciones FK (instrumento, usuario_creador intactos)
    """
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, activo=True)

    if request.method == "POST":
        try:
            calificacion.activo = False
            calificacion.save()

            logger.warning(
                f"Calificacion deleted (logical) - User: {request.user.username}, "
                f"ID: {calificacion.id}, Instrumento: {calificacion.instrumento.codigo_instrumento}"
            )

            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion="DELETE",
                tabla_afectada="CalificacionTributaria",
                registro_id=calificacion.id,
                ip_address=ip_address,
                detalles=f"Calificación eliminada: {calificacion.instrumento.codigo_instrumento}",
            )

            messages.success(request, "Calificación eliminada exitosamente.")
            return redirect("listar_calificaciones")
        except Exception as e:
            logger.error(f"Error deleting calificacion {pk}: {e}", exc_info=True)
            messages.error(request, f"Error al eliminar calificación: {str(e)}")
            return redirect("listar_calificaciones")

    return render(request, "calificaciones/confirmar_eliminar.html", {"objeto": calificacion})


@login_required
@requiere_permiso("crear")
def crear_calificacion_factores(request):
    """
    Crea nueva calificación tributaria usando formulario simplificado de 5 factores.

    Formulario alternativo que calcula el factor total a partir de 5 sub-factores
    individuales (factor_1 a factor_5) proporcionando interfaz más intuitiva para
    el usuario que trabaja con la clasificación SII.

    Parámetros:
        request (HttpRequest): Solicitud HTTP del usuario autenticado.

    Retorna:
        HttpResponse:
            - POST válido: Redirect a 'listar_calificaciones' tras crear registro.
            - POST inválido: Render del formulario con errores.
            - GET: Render de 'calificaciones/form_factores_simple.html'.

    Excepciones:
        Exception: Captura errores de guardado y muestra mensaje al usuario.

    Notas:
        - Requiere permiso: 'crear'
        - Formulario: CalificacionFactoresSimpleForm (5 factores individuales)
        - Cálculo automático: factor_total = suma de 5 factores
        - Registra en LogAuditoria con acción CREATE
        - JSON con tipos de instrumentos para interfaz dinámica
        - Logging implícito vía señales de modelo
    """
    if request.method == "POST":
        form = CalificacionFactoresSimpleForm(request.POST)

        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.usuario_creador = request.user

            try:
                calificacion.save()

                ip_address = obtener_ip_cliente(request)
                LogAuditoria.objects.create(
                    usuario=request.user,
                    accion="CREATE",
                    tabla_afectada="CalificacionTributaria",
                    registro_id=calificacion.id,
                    ip_address=ip_address,
                    detalles=f"Calificación creada con factores: {calificacion.instrumento.codigo_instrumento}",
                )

                messages.success(
                    request,
                    "Calificación creada exitosamente. Factores calculados automáticamente.",
                )
                return redirect("listar_calificaciones")

            except Exception as e:
                messages.error(request, f"Error al guardar la calificación: {str(e)}")
        else:
            messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = CalificacionFactoresSimpleForm()

    instrumentos = InstrumentoFinanciero.objects.filter(activo=True)
    tipos_instrumentos = {inst.id: inst.tipo_instrumento for inst in instrumentos}

    context = {
        "form": form,
        "titulo": "Nueva Calificación con Factores",
        "accion": "Crear",
        "tipos_instrumentos_json": json.dumps(tipos_instrumentos),
    }

    return render(request, "calificaciones/form_factores_simple.html", context)


@login_required
@requiere_permiso("modificar")
def editar_calificacion_factores(request, pk):
    """
    Edita calificación tributaria existente usando formulario de 5 factores.

    Permite modificar una calificación existente mediante el formulario simplificado
    que desglosa el factor total en 5 componentes individuales.

    Parámetros:
        request (HttpRequest): Solicitud HTTP del usuario autenticado.
        pk (int): Clave primaria de la CalificacionTributaria a editar.

    Retorna:
        HttpResponse:
            - POST válido: Redirect a 'listar_calificaciones' tras actualizar.
            - POST inválido: Render del formulario con errores y datos actuales.
            - GET: Render de 'calificaciones/form_factores_simple.html' pre-poblado.

    Excepciones:
        Http404: Si no existe calificación con pk dado o activo=False.

    Notas:
        - Requiere permiso: 'modificar'
        - Solo edita registros activos (activo=True)
        - Formulario: CalificacionFactoresSimpleForm con instance
        - Actualiza usuario_creador al usuario actual
        - Registra en LogAuditoria con acción UPDATE
        - JSON con tipos de instrumentos para interfaz dinámica
        - Context incluye calificación original para comparación
    """
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, activo=True)

    if request.method == "POST":
        form = CalificacionFactoresSimpleForm(request.POST, instance=calificacion)
        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.usuario_creador = request.user
            calificacion.save()

            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion="UPDATE",
                tabla_afectada="CalificacionTributaria",
                registro_id=calificacion.id,
                ip_address=ip_address,
                detalles=f"Calificación editada: {calificacion.instrumento.codigo_instrumento}",
            )

            messages.success(request, "Calificación actualizada exitosamente.")
            return redirect("listar_calificaciones")
    else:
        form = CalificacionFactoresSimpleForm(instance=calificacion)

    instrumentos = InstrumentoFinanciero.objects.filter(activo=True)
    tipos_instrumentos = {inst.id: inst.tipo_instrumento for inst in instrumentos}

    context = {
        "form": form,
        "calificacion": calificacion,
        "titulo": "Editar Calificación",
        "accion": "Actualizar",
        "tipos_instrumentos_json": json.dumps(tipos_instrumentos),
    }
    return render(request, "calificaciones/form_factores_simple.html", context)


# ============================================================================
# SECTION 5: INSTRUMENTOS CRUD OPERATIONS
# ============================================================================
# Functions: listar_instrumentos, crear_instrumento, editar_instrumento,
#            eliminar_instrumento
# Lines: 740-860 (approx. 120 lines)
# ============================================================================


@login_required
@requiere_permiso("consultar")
def listar_instrumentos(request):
    """
    Lista todos los instrumentos financieros activos con búsqueda multi-campo.

    Muestra tabla de instrumentos con capacidad de búsqueda simultánea en código,
    nombre y tipo de instrumento usando operadores OR.

    Parámetros:
        request (HttpRequest): GET request con parámetro opcional:
            - busqueda (str): Término de búsqueda (busca en código, nombre y tipo).

    Retorna:
        HttpResponse: Render de 'calificaciones/listar_instrumentos.html' con
            QuerySet de instrumentos filtrados.

    Notas:
        - Solo muestra instrumentos activos (activo=True)
        - Búsqueda case-insensitive (ICONTAINS)
        - Búsqueda multi-campo: código OR nombre OR tipo
        - Ordenado por codigo_instrumento ascendente
        - Requiere autenticación pero NO requiere permiso específico
    """
    instrumentos = InstrumentoFinanciero.objects.filter(activo=True)

    # Búsqueda
    busqueda = request.GET.get("busqueda")
    if busqueda:
        instrumentos = instrumentos.filter(
            Q(codigo_instrumento__icontains=busqueda)
            | Q(nombre_instrumento__icontains=busqueda)
            | Q(tipo_instrumento__icontains=busqueda)
        )

    instrumentos = instrumentos.order_by("codigo_instrumento")

    return render(
        request, "calificaciones/listar_instrumentos.html", {"instrumentos": instrumentos}
    )


@login_required
@requiere_permiso("crear")
def crear_instrumento(request):
    """
    Crea un nuevo instrumento financiero en el catálogo del sistema.

    Permite agregar instrumentos que luego serán referenciados por las calificaciones
    tributarias. Registra la creación en auditoría.

    Parámetros:
        request (HttpRequest):
            - GET: Muestra formulario vacío
            - POST: Procesa formulario con datos del instrumento

    Retorna:
        HttpResponse:
            - GET: Render de 'calificaciones/form_instrumento.html' con form vacío
            - POST exitoso: Redirect a 'listar_instrumentos' con mensaje de éxito

    Notas:
        - Requiere permiso: 'crear'
        - Campos principales: codigo_instrumento (único), nombre_instrumento, tipo_instrumento
        - Registra acción CREATE en LogAuditoria
        - Logging: INFO en creación exitosa con código y tipo
        - codigo_instrumento debe ser único (constraint en BD)
    """
    if request.method == "POST":
        form = InstrumentoFinancieroForm(request.POST)
        if form.is_valid():
            instrumento = form.save()

            logger.info(
                f"Instrument created - User: {request.user.username}, "
                f"Code: {instrumento.codigo_instrumento}, Type: {instrumento.tipo_instrumento}"
            )

            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion="CREATE",
                tabla_afectada="InstrumentoFinanciero",
                registro_id=instrumento.id,
                ip_address=ip_address,
                detalles=f"Instrumento creado: {instrumento.codigo_instrumento}",
            )

            messages.success(request, "Instrumento creado exitosamente.")
            return redirect("listar_instrumentos")
    else:
        form = InstrumentoFinancieroForm()

    return render(request, "calificaciones/form_instrumento.html", {"form": form})


@login_required
@requiere_permiso("modificar")
def editar_instrumento(request, pk):
    """
    Edita un instrumento financiero existente.

    Permite actualizar información de un instrumento: código, nombre, tipo, emisor
    y descripción. Validación automática de unicidad de código.

    Parámetros:
        request (HttpRequest): Solicitud HTTP del usuario autenticado.
        pk (int): Clave primaria del InstrumentoFinanciero a editar.

    Retorna:
        HttpResponse:
            - POST válido: Redirect a 'listar_instrumentos' tras actualizar.
            - POST inválido: Render del formulario con errores.
            - GET: Render de 'calificaciones/form_instrumento.html' pre-poblado.

    Excepciones:
        Http404: Si no existe instrumento con pk dado o activo=False.

    Notas:
        - Requiere permiso: 'modificar'
        - Solo edita registros activos (activo=True)
        - Formulario: InstrumentoFinancieroForm con instance
        - Validación de unicidad de codigo_instrumento en formulario
        - Registra en LogAuditoria con acción UPDATE
        - No afecta calificaciones asociadas existentes
    """
    instrumento = get_object_or_404(InstrumentoFinanciero, pk=pk, activo=True)

    if request.method == "POST":
        form = InstrumentoFinancieroForm(request.POST, instance=instrumento)
        if form.is_valid():
            instrumento = form.save()

            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion="UPDATE",
                tabla_afectada="InstrumentoFinanciero",
                registro_id=instrumento.id,
                ip_address=ip_address,
                detalles=f"Instrumento editado: {instrumento.codigo_instrumento}",
            )

            messages.success(request, "Instrumento actualizado exitosamente.")
            return redirect("listar_instrumentos")
    else:
        form = InstrumentoFinancieroForm(instance=instrumento)

    context = {"form": form, "instrumento": instrumento}
    return render(request, "calificaciones/form_instrumento.html", context)


@login_required
@requiere_permiso("eliminar")
def eliminar_instrumento(request, pk):
    """
    Realiza eliminación lógica de un instrumento financiero.

    Soft delete que marca activo=False en lugar de eliminar físicamente el registro.
    Protege integridad referencial validando que no existan calificaciones activas
    asociadas antes de permitir la eliminación.

    Parámetros:
        request (HttpRequest): Solicitud HTTP del usuario autenticado.
        pk (int): Clave primaria del InstrumentoFinanciero a eliminar.

    Retorna:
        HttpResponse:
            - POST: Redirect a 'listar_instrumentos' tras eliminar o con error.
            - GET: Render de 'calificaciones/confirmar_eliminar.html' para confirmación.

    Excepciones:
        Http404: Si no existe instrumento con pk dado o activo=False.

    Notas:
        - Requiere permiso: 'eliminar'
        - Validación: Bloquea eliminación si existen CalificacionTributaria activas
        - Eliminación lógica: activo=False (no DELETE físico)
        - Registra en LogAuditoria con acción DELETE
        - Logging: WARNING con usuario y código de instrumento
        - Patrón GET-POST para confirmación de usuario
        - Mensaje de error si tiene calificaciones asociadas
    """
    instrumento = get_object_or_404(InstrumentoFinanciero, pk=pk, activo=True)

    # Verificar si tiene calificaciones asociadas
    tiene_calificaciones = CalificacionTributaria.objects.filter(
        instrumento=instrumento, activo=True
    ).exists()

    if tiene_calificaciones:
        messages.error(
            request, "No se puede eliminar el instrumento porque tiene calificaciones asociadas."
        )
        return redirect("listar_instrumentos")

    if request.method == "POST":
        instrumento.activo = False
        instrumento.save()

        logger.warning(
            f"Instrument deleted (logical) - User: {request.user.username}, "
            f"Code: {instrumento.codigo_instrumento}"
        )

        # Registrar en auditoría
        ip_address = obtener_ip_cliente(request)
        LogAuditoria.objects.create(
            usuario=request.user,
            accion="DELETE",
            tabla_afectada="InstrumentoFinanciero",
            registro_id=instrumento.id,
            ip_address=ip_address,
            detalles=f"Instrumento eliminado: {instrumento.codigo_instrumento}",
        )

        messages.success(request, "Instrumento eliminado exitosamente.")
        return redirect("listar_instrumentos")

    return render(request, "calificaciones/confirmar_eliminar.html", {"objeto": instrumento})


# ============================================================================
# SECTION 6: BULK OPERATIONS
# ============================================================================
# Functions: carga_masiva, exportar_excel, exportar_csv
# Lines: 861-1100 (approx. 240 lines)
# ============================================================================


@login_required
@requiere_permiso("crear")
def carga_masiva(request):
    """
    Procesa carga masiva de calificaciones tributarias desde archivos CSV o Excel.

    Función de alto riesgo que importa múltiples registros en una sola operación.
    Soporta creación automática de instrumentos financieros si no existen y registra
    el resultado completo de la operación en la tabla CargaMasiva.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP con archivo en FILES.

    Retorna:
        HttpResponse:
            - POST: Redirect a 'dashboard' después de procesar.
            - GET: Render de 'calificaciones/carga_masiva.html' con formulario.

    Excepciones:
        ValueError: Si el formato del archivo no es soportado (.xlsx, .csv).
        PermissionError: Si hay problemas de acceso al archivo.
        KeyError: Si faltan campos requeridos en las filas del archivo.

    Notas:
        - Formatos soportados: .xlsx (Excel), .csv (UTF-8)
        - Campos requeridos: codigo_instrumento, fecha_informe
        - Campos opcionales: nombre_instrumento, tipo_instrumento, monto, factor,
          metodo_ingreso, numero_dj, observaciones
        - Estados posibles: EXITOSO (0 errores), PARCIAL (algunos errores), FALLIDO (todos errores)
        - Registra cada fila procesada en CargaMasiva con errores_detalle
        - Logging exhaustivo: INFO (inicio/fin), WARNING (errores por fila), ERROR (crítico)
        - Requiere permiso: @requiere_permiso("crear")
    """
    if request.method == "POST":
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]

            logger.info(
                f"Bulk upload started - User: {request.user.username}, "
                f"File: {archivo.name}, Size: {archivo.size} bytes"
            )

            # Crear registro de carga
            carga = CargaMasiva.objects.create(
                usuario=request.user,
                archivo_nombre=archivo.name,
                archivo=archivo,
                estado="PROCESANDO",
            )

            try:
                # Detectar tipo de archivo
                if archivo.name.endswith(".xlsx"):
                    logger.debug(f"Processing Excel file: {archivo.name}")
                    registros = procesar_excel(archivo)
                elif archivo.name.endswith(".csv"):
                    logger.debug(f"Processing CSV file: {archivo.name}")
                    registros = procesar_csv(archivo)
                else:
                    logger.error(f"Unsupported file format: {archivo.name}")
                    raise ValueError("Formato de archivo no soportado")

                # Procesar registros
                exitosos = 0
                fallidos = 0
                errores = []

                for i, registro in enumerate(registros, start=1):
                    try:
                        # Buscar o crear instrumento
                        instrumento, created = InstrumentoFinanciero.objects.get_or_create(
                            codigo_instrumento=registro["codigo_instrumento"],
                            defaults={
                                "nombre_instrumento": registro.get("nombre_instrumento", ""),
                                "tipo_instrumento": registro.get("tipo_instrumento", "Otro"),
                            },
                        )

                        # Preparar datos de calificación con mapeo dinámico de factores
                        calificacion_data = {
                            'instrumento': instrumento,
                            'usuario_creador': request.user,
                            'monto': Decimal(str(registro["monto"])) if registro.get("monto") else None,
                            'factor': Decimal(str(registro["factor"])) if registro.get("factor") else None,
                            'metodo_ingreso': registro.get("metodo_ingreso", "MONTO"),
                            'numero_dj': registro.get("numero_dj", ""),
                            'fecha_informe': registro["fecha_informe"],
                            'observaciones': registro.get("observaciones", ""),
                            # Campos metadata administrativos (TASK-004 - HDU_Inacap.xlsx)
                            'secuencia': int(registro["secuencia"]) if registro.get("secuencia") else 0,
                            'numero_dividendo': int(registro["numero_dividendo"]) if registro.get("numero_dividendo") else 0,
                            'tipo_sociedad': registro.get("tipo_sociedad", None),
                            'valor_historico': Decimal(str(registro["valor_historico"])) if registro.get("valor_historico") else Decimal('0'),
                            'mercado': registro.get("mercado", None),
                            'ejercicio': int(registro["ejercicio"]) if registro.get("ejercicio") else 0,
                        }
                        
                        # Mapear dinámicamente todos los factores (8-37) desde el archivo
                        for factor_num in range(8, 38):
                            factor_key = f'factor_{factor_num}'
                            if factor_key in registro and registro[factor_key]:
                                try:
                                    calificacion_data[factor_key] = Decimal(str(registro[factor_key]))
                                except (ValueError, TypeError):
                                    # Si el valor no es válido, dejar en None (default=0 en modelo)
                                    calificacion_data[factor_key] = None
                        
                        # Crear calificación con todos los datos
                        CalificacionTributaria.objects.create(**calificacion_data)

                        exitosos += 1
                    except ValidationError as e:
                        # Captura errores de validación del modelo (ej: suma de factores > 1)
                        fallidos += 1
                        error_msg = e.message if hasattr(e, 'message') else str(e)
                        errores.append(f"Fila {i}: {error_msg}")
                        logger.warning(f"Bulk upload row {i} validation error: {e}")
                    except IntegrityError as e:
                        # Captura errores de integridad de base de datos
                        fallidos += 1
                        error_str = str(e).lower()
                        # Detectar si es un error de duplicado por unique_together
                        if 'unique' in error_str or 'duplicate' in error_str or 'already exists' in error_str:
                            errores.append(f"Fila {i}: Error - Este registro ya existe en el sistema (Duplicado).")
                            logger.warning(f"Bulk upload row {i} duplicate record: {e}")
                        else:
                            errores.append(f"Fila {i}: Error de integridad de datos - {str(e)}")
                            logger.warning(f"Bulk upload row {i} integrity error: {e}")
                    except KeyError as e:
                        fallidos += 1
                        errores.append(f"Fila {i}: Campo requerido faltante - {str(e)}")
                        logger.warning(f"Bulk upload row {i} missing field: {e}")
                    except ValueError as e:
                        fallidos += 1
                        errores.append(f"Fila {i}: Valor inválido - {str(e)}")
                        logger.warning(f"Bulk upload row {i} invalid value: {e}")
                    except Exception as e:
                        # Captura cualquier otro error inesperado
                        fallidos += 1
                        errores.append(f"Fila {i}: {str(e)}")
                        logger.error(f"Bulk upload row {i} unexpected error: {e}", exc_info=True)

                # Actualizar registro de carga
                carga.registros_procesados = len(registros)
                carga.registros_exitosos = exitosos
                carga.registros_fallidos = fallidos
                carga.errores_detalle = "\n".join(errores)

                if fallidos == 0:
                    carga.estado = "EXITOSO"
                    logger.info(
                        f"Bulk upload completed successfully - User: {request.user.username}, "
                        f"File: {archivo.name}, Records: {exitosos}/{len(registros)}"
                    )
                elif exitosos > 0:
                    carga.estado = "PARCIAL"
                    logger.warning(
                        f"Bulk upload partially failed - User: {request.user.username}, "
                        f"File: {archivo.name}, Success: {exitosos}, Failed: {fallidos}"
                    )
                else:
                    carga.estado = "FALLIDO"
                    logger.error(
                        f"Bulk upload failed completely - User: {request.user.username}, "
                        f"File: {archivo.name}, Total records: {len(registros)}"
                    )

                carga.save()

                # Registrar en auditoría
                ip_address = obtener_ip_cliente(request)
                LogAuditoria.objects.create(
                    usuario=request.user,
                    accion="CREATE",
                    tabla_afectada="CargaMasiva",
                    registro_id=carga.id,
                    ip_address=ip_address,
                    detalles=f"Carga masiva: {exitosos} exitosos, {fallidos} fallidos",
                )

                messages.success(
                    request,
                    f"Procesados {exitosos} registros correctamente. {fallidos} con errores.",
                )

            except ValueError as e:
                logger.error(f"File format error - File: {archivo.name}, Error: {str(e)}")
                carga.estado = "FALLIDO"
                carga.errores_detalle = f"Error de formato: {str(e)}"
                carga.save()
                messages.error(request, f"Error al procesar archivo: Formato inválido - {str(e)}")
            except PermissionError as e:
                logger.error(f"File access error - File: {archivo.name}, Error: {str(e)}")
                carga.estado = "FALLIDO"
                carga.errores_detalle = f"Error de acceso al archivo: {str(e)}"
                carga.save()
                messages.error(request, f"Error al acceder al archivo: {str(e)}")
            except Exception as e:
                logger.error(
                    f"Critical error in bulk upload - User: {request.user.username}, "
                    f"File: {archivo.name}, Error: {str(e)}",
                    exc_info=True,
                )
                carga.estado = "FALLIDO"
                carga.errores_detalle = str(e)
                carga.save()
                messages.error(request, f"Error inesperado al procesar archivo: {str(e)}")

            return redirect("dashboard")
    else:
        form = CargaMasivaForm()

    # Obtener historial de cargas recientes para mostrar en la plantilla
    cargas_anteriores = CargaMasiva.objects.filter(usuario=request.user).order_by('-fecha_carga')[:10]

    return render(request, "calificaciones/carga_masiva.html", {
        "form": form,
        "cargas_anteriores": cargas_anteriores
    })


@login_required
@requiere_permiso("crear")
def descargar_plantilla(request, formato='xlsx'):
    """
    Genera y descarga una plantilla para carga masiva de calificaciones en formato Excel o CSV.
    
    Crea un archivo con las columnas requeridas y una fila de ejemplo
    para facilitar la carga masiva de calificaciones tributarias.
    
    Parámetros:
        request: HttpRequest
        formato: str - 'xlsx' o 'csv' (default: 'xlsx')
    
    Retorna:
        HttpResponse: Archivo descargable con:
            - Excel: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
            - CSV: text/csv; charset=utf-8
            - Columnas: metadata (8) + 30 factores (8-37)
    """
    # Encabezados de metadata
    headers = [
        "codigo_instrumento",
        "fecha_informe",
        "mercado",
        "tipo_sociedad",
        "secuencia",
        "numero_dividendo",
        "ejercicio",
        "numero_dj",
    ]
    
    # Agregar headers de 30 factores
    for i in range(8, 38):
        headers.append(f"factor_{i}")
    
    # Fila de ejemplo
    ejemplo = [
        "INST001",  # codigo_instrumento
        "2025-01-15",  # fecha_informe (YYYY-MM-DD)
        "ACN",  # mercado
        "A",  # tipo_sociedad
        "1",  # secuencia
        "0",  # numero_dividendo
        "2025",  # ejercicio
        "DJ-2025-001",  # numero_dj
    ]
    
    # Agregar valores de ejemplo para factores (0.0333... para que sumen 1.0)
    for i in range(30):
        ejemplo.append("0.0333")
    
    if formato == 'csv':
        # Generar CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename=plantilla_carga_masiva.csv'
        
        # Agregar BOM para Excel en Windows
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerow(ejemplo)
        
        logger.info(f"CSV template downloaded - User: {request.user.username}")
    else:
        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Carga Masiva"
        
        ws.append(headers)
        ws.append(ejemplo)
        
        # Estilo de encabezados
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="F37021", end_color="F37021", fill_type="solid")
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=plantilla_carga_masiva.xlsx"
        
        wb.save(response)
        
        logger.info(f"Excel template downloaded - User: {request.user.username}")
    
    return response


@login_required
@requiere_permiso("consultar")
def exportar_excel(request):
    """
    Exporta calificaciones a formato Excel (.xlsx) con todos los 30 factores.

    Genera archivo Excel con todas las calificaciones tributarias aplicando los mismos
    filtros que la vista de listado. Incluye metadata completa y 30 factores (8-37).

    Parámetros:
        request (HttpRequest): Solicitud HTTP con filtros opcionales en GET:
            - codigo_instrumento: Búsqueda parcial por código/nombre
            - mercado: ACN, CFI, FFM
            - tipo_sociedad: A (Corredora), C (Bolsa)
            - ejercicio: Año (int)
            - numero_dj: Número de DJ

    Retorna:
        HttpResponse: Archivo Excel descargable con:
            - Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
            - Filename: calificaciones_YYYYMMDD_HHMMSS.xlsx
            - Columnas: 41+ campos (ID, Instrumento, Metadata, 30 Factores, Observaciones)

    Notas:
        - Requiere permiso: 'consultar'
        - Librería: openpyxl
        - Solo exporta registros activos (activo=True)
        - Aplica mismos filtros que listar_calificaciones
        - Query optimizado con select_related('instrumento', 'usuario_creador')
    """
    # Aplicar filtros (misma lógica que listar_calificaciones)
    calificaciones = CalificacionTributaria.objects.filter(activo=True).select_related(
        "instrumento", "usuario_creador"
    )

    # FILTROS PRINCIPALES
    mercado = request.GET.get("mercado", "").strip()
    tipo_sociedad = request.GET.get("tipo_sociedad", "").strip()
    ejercicio = request.GET.get("ejercicio", "").strip()
    codigo_instrumento = request.GET.get("codigo_instrumento", "").strip()
    numero_dj = request.GET.get("numero_dj", "").strip()

    if mercado:
        calificaciones = calificaciones.filter(mercado__iexact=mercado)
    
    if tipo_sociedad:
        calificaciones = calificaciones.filter(tipo_sociedad__iexact=tipo_sociedad)
    
    if ejercicio:
        try:
            ejercicio_int = int(ejercicio)
            calificaciones = calificaciones.filter(ejercicio=ejercicio_int)
        except ValueError:
            pass
    
    if codigo_instrumento:
        calificaciones = calificaciones.filter(
            Q(instrumento__codigo_instrumento__icontains=codigo_instrumento) |
            Q(instrumento__nombre_instrumento__icontains=codigo_instrumento)
        )
    
    if numero_dj:
        calificaciones = calificaciones.filter(numero_dj__icontains=numero_dj)

    logger.info(
        f"Excel export - User: {request.user.username}, Records: {calificaciones.count()}, "
        f"Filters: mercado={mercado}, tipo_sociedad={tipo_sociedad}, ejercicio={ejercicio}"
    )

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calificaciones"

    # Encabezados dinámicos con 30 factores
    headers = [
        "ID",
        "Código Instrumento",
        "Nombre Instrumento",
        "Fecha Informe",
        "Mercado",
        "Secuencia",
        "Origen (Tipo Soc)",
        "N° DJ",
        "Ejercicio",
        "Valor Histórico",
        "Monto",
    ]
    
    # Agregar headers de 30 factores dinámicamente
    for i in range(8, 38):
        headers.append(f"Factor {i}")
    
    headers.extend([
        "Método Ingreso",
        "Usuario Creador",
        "Fecha Creación",
        "Observaciones"
    ])
    
    ws.append(headers)

    # Datos
    for cal in calificaciones:
        row = [
            cal.id,
            cal.instrumento.codigo_instrumento,
            cal.instrumento.nombre_instrumento,
            cal.fecha_informe.strftime("%Y-%m-%d") if cal.fecha_informe else "",
            cal.mercado or "",
            cal.secuencia if cal.secuencia else "",
            cal.tipo_sociedad or "",
            cal.numero_dj,
            cal.ejercicio if cal.ejercicio else "",
            float(cal.valor_historico) if cal.valor_historico else None,
            float(cal.monto) if cal.monto else None,
        ]
        
        # Agregar 30 factores dinámicamente
        for i in range(8, 38):
            factor_value = getattr(cal, f"factor_{i}", None)
            row.append(float(factor_value) if factor_value else None)
        
        row.extend([
            cal.metodo_ingreso or "",
            cal.usuario_creador.username if cal.usuario_creador else "",
            cal.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S") if cal.fecha_creacion else "",
            cal.observaciones or "",
        ])
        
        ws.append(row)

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename=calificaciones_{timestamp}.xlsx'

    wb.save(response)

    # Registrar en auditoría
    ip_address = obtener_ip_cliente(request)
    LogAuditoria.objects.create(
        usuario=request.user,
        accion="READ",
        tabla_afectada="CalificacionTributaria",
        ip_address=ip_address,
        detalles=f"Exportación Excel: {calificaciones.count()} registros con filtros aplicados",
    )

    return response


@login_required
@requiere_permiso("consultar")
def exportar_csv(request):
    """
    Exporta calificaciones a formato CSV con todos los 30 factores.

    Genera archivo CSV compatible con Excel y otras herramientas. Incluye metadata
    completa y 30 factores (8-37). Aplica los mismos filtros que la vista de listado.

    Parámetros:
        request (HttpRequest): Solicitud HTTP con filtros opcionales en GET:
            - codigo_instrumento: Búsqueda parcial por código/nombre
            - mercado: ACN, CFI, FFM
            - tipo_sociedad: A (Corredora), C (Bolsa)
            - ejercicio: Año (int)
            - numero_dj: Número de DJ

    Retorna:
        HttpResponse: Archivo CSV descargable con:
            - Content-Type: text/csv
            - Encoding: UTF-8 (compatible con caracteres especiales)
            - Filename: calificaciones_YYYYMMDD_HHMMSS.csv
            - Columnas: 41+ campos (ID, Instrumento, Metadata, 30 Factores, Observaciones)

    Notas:
        - Requiere permiso: 'consultar'
        - Librería: csv (stdlib)
        - Solo exporta registros activos (activo=True)
        - Aplica mismos filtros que listar_calificaciones
        - Query optimizado con select_related('instrumento', 'usuario_creador')
        - Separador: coma (,)
    """
    # Aplicar filtros (misma lógica que listar_calificaciones)
    calificaciones = CalificacionTributaria.objects.filter(activo=True).select_related(
        "instrumento", "usuario_creador"
    )

    # FILTROS PRINCIPALES
    mercado = request.GET.get("mercado", "").strip()
    tipo_sociedad = request.GET.get("tipo_sociedad", "").strip()
    ejercicio = request.GET.get("ejercicio", "").strip()
    codigo_instrumento = request.GET.get("codigo_instrumento", "").strip()
    numero_dj = request.GET.get("numero_dj", "").strip()

    if mercado:
        calificaciones = calificaciones.filter(mercado__iexact=mercado)
    
    if tipo_sociedad:
        calificaciones = calificaciones.filter(tipo_sociedad__iexact=tipo_sociedad)
    
    if ejercicio:
        try:
            ejercicio_int = int(ejercicio)
            calificaciones = calificaciones.filter(ejercicio=ejercicio_int)
        except ValueError:
            pass
    
    if codigo_instrumento:
        calificaciones = calificaciones.filter(
            Q(instrumento__codigo_instrumento__icontains=codigo_instrumento) |
            Q(instrumento__nombre_instrumento__icontains=codigo_instrumento)
        )
    
    if numero_dj:
        calificaciones = calificaciones.filter(numero_dj__icontains=numero_dj)

    logger.info(
        f"CSV export - User: {request.user.username}, Records: {calificaciones.count()}, "
        f"Filters: mercado={mercado}, tipo_sociedad={tipo_sociedad}, ejercicio={ejercicio}"
    )

    response = HttpResponse(content_type="text/csv")
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename=calificaciones_{timestamp}.csv'

    writer = csv.writer(response)
    
    # Encabezados dinámicos con 30 factores
    headers = [
        "ID",
        "Código Instrumento",
        "Nombre Instrumento",
        "Fecha Informe",
        "Mercado",
        "Secuencia",
        "Origen (Tipo Soc)",
        "N° DJ",
        "Ejercicio",
        "Valor Histórico",
        "Monto",
    ]
    
    # Agregar headers de 30 factores dinámicamente
    for i in range(8, 38):
        headers.append(f"Factor {i}")
    
    headers.extend([
        "Método Ingreso",
        "Usuario Creador",
        "Fecha Creación",
        "Observaciones"
    ])
    
    writer.writerow(headers)

    # Datos
    for cal in calificaciones:
        row = [
            cal.id,
            cal.instrumento.codigo_instrumento,
            cal.instrumento.nombre_instrumento,
            cal.fecha_informe.strftime("%Y-%m-%d") if cal.fecha_informe else "",
            cal.mercado or "",
            cal.secuencia if cal.secuencia else "",
            cal.tipo_sociedad or "",
            cal.numero_dj,
            cal.ejercicio if cal.ejercicio else "",
            cal.valor_historico if cal.valor_historico else "",
            cal.monto if cal.monto else "",
        ]
        
        # Agregar 30 factores dinámicamente
        for i in range(8, 38):
            factor_value = getattr(cal, f"factor_{i}", None)
            row.append(factor_value if factor_value else "")
        
        row.extend([
            cal.metodo_ingreso or "",
            cal.usuario_creador.username if cal.usuario_creador else "",
            cal.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S") if cal.fecha_creacion else "",
            cal.observaciones or "",
        ])
        
        writer.writerow(row)

    # Registrar en auditoría
    ip_address = obtener_ip_cliente(request)
    LogAuditoria.objects.create(
        usuario=request.user,
        accion="READ",
        tabla_afectada="CalificacionTributaria",
        ip_address=ip_address,
        detalles=f"Exportación CSV: {calificaciones.count()} registros con filtros aplicados",
    )

    return response


# ============================================================================
# SECTION 7: USER MANAGEMENT
# ============================================================================
# Functions: mi_perfil, registro, admin_gestionar_usuarios,
#            desbloquear_cuenta_manual, ver_historial_login_usuario
# Lines: 1101-1300 (approx. 200 lines)
# ============================================================================


@login_required
def mi_perfil(request):
    """
    Muestra y permite editar el perfil del usuario autenticado.

    Presenta formulario para actualizar información personal: nombre, apellido,
    email, teléfono y departamento. Crea perfil automáticamente si no existe.

    Parámetros:
        request (HttpRequest):
            - GET: Muestra formulario con datos actuales
            - POST: Actualiza información del perfil

    Retorna:
        HttpResponse:
            - GET: Render de 'calificaciones/mi_perfil.html' con datos del perfil
            - POST: Redirect a 'mi_perfil' tras actualización exitosa

    Notas:
        - Requiere autenticación: @login_required
        - No requiere permiso específico (todos pueden editar su perfil)
        - Crea PerfilUsuario automáticamente si no existe
        - Actualiza User (first_name, last_name, email) y PerfilUsuario (telefono, departamento)
        - No permite cambiar username o password (requiere vistas específicas)
    """
    try:
        perfil = request.user.perfilusuario
    except PerfilUsuario.DoesNotExist:
        # Crear perfil si no existe
        perfil = PerfilUsuario.objects.create(usuario=request.user)

    if request.method == "POST":
        # Actualizar información básica
        request.user.first_name = request.POST.get("first_name", "")
        request.user.last_name = request.POST.get("last_name", "")
        request.user.email = request.POST.get("email", "")
        request.user.save()

        # Actualizar perfil
        perfil.telefono = request.POST.get("telefono", "")
        perfil.departamento = request.POST.get("departamento", "")
        perfil.save()

        messages.success(request, "Perfil actualizado exitosamente.")
        return redirect("mi_perfil")

    context = {
        "perfil": perfil,
    }

    return render(request, "calificaciones/mi_perfil.html", context)


def registro(request):
    """
    Permite el auto-registro de nuevos usuarios en el sistema.

    Función pública (no requiere login) que permite crear cuentas de usuario con
    asignación automática del rol 'Auditor' por defecto. Administradores pueden
    cambiar el rol posteriormente.

    Parámetros:
        request (HttpRequest): Solicitud HTTP (pública, sin autenticación).

    Retorna:
        HttpResponse:
            - POST válido: Redirect a 'login' con mensaje de éxito.
            - POST inválido: Render del formulario con errores.
            - GET: Render de 'calificaciones/registro.html' con formulario vacío.

    Notas:
        - No requiere login (vista pública)
        - No requiere permisos especiales
        - Formulario: RegistroForm (hereda de UserCreationForm)
        - Rol por defecto: 'Auditor' (menor privilegio)
        - Crea automáticamente PerfilUsuario asociado
        - Registra en LogAuditoria con acción CREATE
        - Usuario nuevo debe hacer login después de registro
        - Validaciones: username único, password strength, email válido
    """
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Crear perfil de usuario con rol por defecto (Auditor)
            rol_auditor = Rol.objects.get(nombre_rol="Auditor")
            PerfilUsuario.objects.create(usuario=user, rol=rol_auditor)

            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=user,
                accion="CREATE",
                tabla_afectada="User",
                registro_id=user.id,
                ip_address=ip_address,
                detalles=f"Usuario registrado: {user.username}",
            )

            messages.success(request, "Registro exitoso. Ya puedes iniciar sesión.")
            return redirect("login")
    else:
        form = RegistroForm()

    return render(request, "calificaciones/registro.html", {"form": form})


@login_required
@requiere_permiso("admin")
def admin_gestionar_usuarios(request):
    """
    Panel administrativo de gestión completa de usuarios del sistema.

    Vista exclusiva para administradores que muestra listado de todos los usuarios
    con información detallada de seguridad: cuentas bloqueadas, intentos de login
    recientes (últimos 7 días), intentos fallidos, y rol asignado.

    Parámetros:
        request (HttpRequest): Solicitud del administrador autenticado.

    Retorna:
        HttpResponse: Render de 'calificaciones/admin/gestionar_usuarios.html' con:
            - usuarios: QuerySet de User con perfiles, bloqueos e intentos
            - total_usuarios: Conteo total de usuarios
            - usuarios_bloqueados: Conteo de cuentas actualmente bloqueadas

    Notas:
        - Requiere permiso: 'admin' (solo Administrador)
        - Query optimizado: select_related('perfilusuario__rol')
        - Agrega atributos dinámicos a cada usuario:
            - cuenta_bloqueada: Instancia de CuentaBloqueada si existe
            - intentos_recientes: Conteo de intentos últimos 7 días
            - intentos_fallidos_recientes: Conteo de fallos últimos 7 días
        - Constante: RECENT_ACTIVITY_DAYS = 7
        - Logging: DEBUG con username del admin
        - Desde esta vista admin puede desbloquear cuentas
    """
    logger.debug(f"Admin user management accessed - Admin: {request.user.username}")

    # Obtener todos los usuarios con sus perfiles
    usuarios = User.objects.all().select_related("perfilusuario__rol").order_by("username")

    # Agregar información adicional a cada usuario
    for usuario in usuarios:
        # Verificar si tiene cuenta bloqueada
        usuario.cuenta_bloqueada = CuentaBloqueada.objects.filter(
            usuario=usuario, bloqueada=True
        ).first()

        # Contar intentos de login recientes
        fecha_limite = timezone.now() - timedelta(days=RECENT_ACTIVITY_DAYS)
        usuario.intentos_recientes = IntentoLogin.objects.filter(
            username=usuario.username, fecha_hora__gte=fecha_limite
        ).count()

        # Contar intentos fallidos recientes
        usuario.intentos_fallidos_recientes = IntentoLogin.objects.filter(
            username=usuario.username, exitoso=False, fecha_hora__gte=fecha_limite
        ).count()

    context = {
        "usuarios": usuarios,
        "total_usuarios": usuarios.count(),
        "usuarios_bloqueados": sum(
            1 for u in usuarios if hasattr(u, "cuenta_bloqueada") and u.cuenta_bloqueada
        ),
    }

    return render(request, "calificaciones/admin/gestionar_usuarios.html", context)


@login_required
@requiere_permiso("admin")
def desbloquear_cuenta_manual(request, user_id):
    """
    Desbloquea manualmente una cuenta de usuario bloqueada por intentos fallidos.

    Permite a administradores desbloquear cuentas sin esperar los 30 minutos de
    bloqueo automático. Registra la acción en auditoría con identificación del admin.

    Parámetros:
        request (HttpRequest): Solicitud del administrador autenticado.
        user_id (int): ID del usuario cuya cuenta se desea desbloquear.

    Retorna:
        HttpResponse: Redirect a 'admin_gestionar_usuarios' con mensaje de resultado.

    Excepciones:
        Http404: Si el usuario no existe.

    Notas:
        - Requiere permiso: 'admin'
        - Registra acción ACCOUNT_UNLOCKED en LogAuditoria
        - Logging: WARNING con admin y target user identificados
        - Marca bloqueada=False y actualiza fecha_desbloqueo
        - Muestra warning si la cuenta no estaba bloqueada
        - Acción crítica requerida por feedback del profesor
    """
    # Obtener el usuario a desbloquear
    user = get_object_or_404(User, id=user_id)

    # Buscar cuenta bloqueada
    cuenta_bloqueada = CuentaBloqueada.objects.filter(usuario=user, bloqueada=True).first()

    if not cuenta_bloqueada:
        messages.warning(request, f"La cuenta de {user.username} no está bloqueada.")
        return redirect("admin_gestionar_usuarios")

    # Desbloquear la cuenta
    cuenta_bloqueada.bloqueada = False
    cuenta_bloqueada.fecha_desbloqueo = timezone.now()
    cuenta_bloqueada.save()

    logger.warning(
        f"Account manually unlocked - Admin: {request.user.username}, "
        f"Target user: {user.username}, Previous attempts: {cuenta_bloqueada.intentos_fallidos}"
    )

    # Registrar en auditoría (CRÍTICO - Feedback del profesor)
    ip_address = obtener_ip_cliente(request)
    LogAuditoria.objects.create(
        usuario=request.user,  # Admin que desbloqueó
        accion="ACCOUNT_UNLOCKED",
        tabla_afectada="CuentaBloqueada",
        registro_id=cuenta_bloqueada.id,
        ip_address=ip_address,
        detalles=f"Cuenta de {user.username} desbloqueada manualmente por {request.user.username}",
    )

    messages.success(request, f"✅ Cuenta de {user.username} desbloqueada exitosamente.")

    return redirect("admin_gestionar_usuarios")


@login_required
@requiere_permiso("admin")
def ver_historial_login_usuario(request, user_id):
    """
    Muestra historial detallado de intentos de login de un usuario específico.

    Vista administrativa que proporciona trazabilidad completa de actividad de
    autenticación de un usuario: intentos exitosos, fallidos, bloqueos y desbloqueos.
    Útil para auditoría de seguridad y troubleshooting.

    Parámetros:
        request (HttpRequest): Solicitud del administrador autenticado.
        user_id (int): ID del usuario cuyo historial se desea consultar.

    Retorna:
        HttpResponse: Render de 'calificaciones/admin/historial_login.html' con:
            - usuario: Instancia del User consultado
            - intentos: Últimos 50 IntentoLogin ordenados por fecha desc
            - logs_login: Últimos 50 LogAuditoria de acciones LOGIN/LOGOUT/LOCK/UNLOCK

    Excepciones:
        Http404: Si no existe usuario con user_id dado.

    Notas:
        - Requiere permiso: 'admin'
        - Límite: 50 registros más recientes de cada tipo
        - IntentoLogin: incluye exitoso, fecha_hora, ip_address
        - LogAuditoria: filtrado por acciones LOGIN, LOGOUT, ACCOUNT_LOCKED, ACCOUNT_UNLOCKED
        - Ordenamiento: descendente por fecha_hora (más recientes primero)
        - Útil para investigar problemas de acceso de usuarios
    """
    user = get_object_or_404(User, id=user_id)

    # Obtener todos los intentos de login del usuario
    intentos = IntentoLogin.objects.filter(username=user.username).order_by("-fecha_hora")[
        :50
    ]  # Últimos 50 intentos

    # Obtener logs de auditoría relacionados con login
    logs_login = LogAuditoria.objects.filter(
        usuario=user, accion__in=["LOGIN", "LOGOUT", "ACCOUNT_LOCKED", "ACCOUNT_UNLOCKED"]
    ).order_by("-fecha_hora")[:50]

    context = {
        "usuario": user,
        "intentos": intentos,
        "logs_login": logs_login,
    }

    return render(request, "calificaciones/admin/historial_login.html", context)


# ============================================================================
# SECTION 8: AUDITING AND COMPLIANCE
# ============================================================================
# Functions: registro_auditoria
# Lines: 1301-1350 (approx. 50 lines)
# ============================================================================


@login_required
@requiere_permiso("consultar")
def registro_auditoria(request):
    """
    Muestra el registro completo de auditoría del sistema con capacidad de filtrado.

    Vista exclusiva para Administradores y Auditores que permite revisar todas las
    acciones registradas en LogAuditoria con filtros por usuario, acción y fechas.

    Parámetros:
        request (HttpRequest): GET request con parámetros opcionales:
            - usuario (int): ID del usuario para filtrar logs.
            - accion (str): Tipo de acción (LOGIN, LOGOUT, CREATE, UPDATE, DELETE, etc.).
            - fecha_desde (str): Fecha mínima (formato YYYY-MM-DD).
            - fecha_hasta (str): Fecha máxima (formato YYYY-MM-DD).

    Retorna:
        HttpResponse: Render de 'calificaciones/registro_auditoria.html' con:
            - logs: QuerySet de LogAuditoria filtrado y ordenado
            - usuarios: Lista de todos los usuarios para filtro
            - acciones: Lista de acciones únicas para filtro
            - Parámetros de filtro en context

    Notas:
        - Requiere permiso: 'admin' (Administrador o Auditor)
        - Ordenado por fecha_hora descendente (más recientes primero)
        - Query optimizado con select_related('usuario')
        - Limitado a MAX_AUDIT_LOG_RECORDS registros por defecto
        - Acciones incluyen: LOGIN, LOGOUT, CREATE, UPDATE, DELETE, ACCOUNT_LOCKED, ACCOUNT_UNLOCKED
    """
    logs = LogAuditoria.objects.all().select_related("usuario").order_by("-fecha_hora")

    # Filtros
    usuario_id = request.GET.get("usuario")
    accion = request.GET.get("accion")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if usuario_id:
        logs = logs.filter(usuario_id=usuario_id)

    if accion:
        logs = logs.filter(accion=accion)

    if fecha_desde:
        logs = logs.filter(fecha_hora__date__gte=fecha_desde)

    if fecha_hasta:
        logs = logs.filter(fecha_hora__date__lte=fecha_hasta)

    # Limitar registros para performance
    logs = logs[:MAX_AUDIT_LOG_RECORDS]

    # Lista de usuarios para filtro
    usuarios = User.objects.filter(is_active=True).order_by("username")

    # ✅ Lista de acciones disponibles
    acciones = LogAuditoria.ACCIONES

    context = {
        "logs": logs,
        "usuarios": usuarios,
        "acciones": acciones,
        "usuario_filtro": usuario_id,
        "accion_filtro": accion,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }

    return render(request, "calificaciones/registro_auditoria.html", context)


# ============================================================================
# SECTION 9: API ENDPOINTS AND MISCELLANEOUS
# ============================================================================
# Functions: calcular_factores_ajax, home
# Lines: 1351-1480 (approx. 130 lines)
# ============================================================================


def calcular_factores_ajax(request):
    """
    API endpoint para calcular automáticamente factores proporcionales desde montos vía AJAX.

    Recibe 30 montos (monto_8 a monto_37), calcula la suma total y determina el factor
    proporcional de cada uno como: factor_i = monto_i / suma_total. Valida REGLA B
    (suma de factores 8-16 <= 1.0).

    Parámetros:
        request (HttpRequest): GET/POST request con parámetros:
            - monto_8 a monto_37 (str, optional): Montos para cada factor. Por defecto: "0".

    Retorna:
        JsonResponse: JSON con estructura:
            {
                "success": true,
                "factores": {
                    "factor_8": "0.25000000",
                    "factor_9": "0.25000000",
                    ...
                    "factor_37": "0.00000000"
                },
                "suma_montos": "1000.00",
                "suma_factores_criticos": "0.95000000",
                "es_valido": true,
                "mensaje_error": "",
                "nombres": {"factor_8": "Factor 8", ...}
            }

        JsonResponse (error): Si falla, retorna {"success": false, "error": mensaje} con status 400.

    Notas:
        - Soporta 30 factores completos (8-37)
        - Precisión: 8 decimales para factores
        - Valida REGLA B: suma factores 8-16 <= 1
        - Conversión automática de valores inválidos a Decimal("0")
        - Utilizado por formulario de calificación completo
        - No requiere autenticación (público)
    """
    try:
        logger.debug(
            f"Factor calculation API called - User: {request.user.username if request.user.is_authenticated else 'Anonymous'}"
        )

        # Soportar tanto GET como POST
        params = request.GET if request.method == 'GET' else request.POST

        # Extraer TODOS los montos (8-37)
        montos = {}
        for i in range(8, 38):
            monto_key = f"monto_{i}"
            montos[monto_key] = params.get(monto_key, "0")

        # Convertir a Decimal
        montos_decimal = {}
        for key, value in montos.items():
            try:
                montos_decimal[key] = Decimal(value) if value else Decimal("0")
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid decimal value for {key}: {value}, Error: {e}")
                montos_decimal[key] = Decimal("0")

        # Calcular suma total de TODOS los montos
        suma_montos = sum(montos_decimal.values())

        # Calcular factores proporcionalmente
        factores = {}
        for key, monto in montos_decimal.items():
            factor_key = key.replace("monto", "factor")
            if suma_montos > 0:
                factor_value = monto / suma_montos
                # Redondear a 8 decimales
                factores[factor_key] = str(factor_value.quantize(Decimal("0.00000001")))
            else:
                factores[factor_key] = "0.00000000"

        # Calcular suma de factores CRÍTICOS (8-16) para REGLA B
        suma_factores_criticos = sum([
            Decimal(factores[f"factor_{i}"]) for i in range(8, 17)
        ])

        # Validar REGLA B
        es_valido = suma_factores_criticos <= Decimal("1.0")
        mensaje_error = ""
        if not es_valido:
            mensaje_error = (
                f"La suma de factores 8-16 es {suma_factores_criticos:.8f}, "
                f"debe ser ≤ 1.00000000 (REGLA B)"
            )

        # Generar nombres de factores
        nombres = {
            'factor_8': 'Con crédito por IDPC generados a contar del 01.01.2017',
            'factor_9': 'Con crédito por IDPC generados hasta el 31.12.2016',
        }
        for i in range(10, 38):
            nombres[f'factor_{i}'] = f'Factor {i}'

        return JsonResponse(
            {
                "success": True,
                "factores": factores,
                "suma_montos": str(suma_montos),
                "suma_factores_criticos": str(suma_factores_criticos),
                "es_valido": es_valido,
                "mensaje_error": mensaje_error,
                "nombres": nombres,
            }
        )

    except Exception as e:
        logger.error(f"Factor calculation error - Error: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "error": str(e)}, status=400)


def home(request):
    """
    Vista para la página de inicio del sistema NUAM.

    Página de aterrizaje pública que muestra información general del sistema.
    No requiere autenticación.

    Parámetros:
        request (HttpRequest): Solicitud HTTP (autenticada o anónima).

    Retorna:
        HttpResponse: Render de 'home.html' con página de inicio.

    Notas:
        - No requiere autenticación ni permisos
        - Template: home.html
        - Redirige aquí tras logout exitoso
        - Puede incluir enlaces a login, información del sistema, etc.
    """
    return render(request, "home.html")
