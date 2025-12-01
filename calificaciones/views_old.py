from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
import csv
import io

from .models import (
    CalificacionTributaria, 
    InstrumentoFinanciero, 
    LogAuditoria, 
    CargaMasiva,
    Rol, 
    PerfilUsuario,
    IntentoLogin,  # Se agregó para auditoría de login
    CuentaBloqueada  # Se agregó para manejo de cuentas bloqueadas
)
from .forms import (
    CalificacionTributariaForm, 
    InstrumentoFinancieroForm, 
    CargaMasivaForm,
    RegistroForm
)
from .permissions import requiere_permiso


# ==========================================
# Nueva función: Obtener IP del cliente
# ==========================================
def obtener_ip_cliente(request):
    """Obtiene la IP real del cliente considerando proxies"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ==========================================
# Nueva función: Verificar si la cuenta está bloqueada
# ==========================================
def verificar_cuenta_bloqueada(username):
    """
    Verifica si una cuenta está bloqueada
    Retorna: (bloqueada: bool, mensaje: str, minutos_restantes: int)
    """
    try:
        user = User.objects.get(username=username)
        cuenta_bloqueada = CuentaBloqueada.objects.filter(usuario=user, bloqueada=True).first()
        
        if cuenta_bloqueada:
            # Verificar si ya pasó el tiempo de bloqueo (30 minutos)
            tiempo_bloqueo = timezone.now() - cuenta_bloqueada.fecha_bloqueo
            MINUTOS_BLOQUEO = 30
            
            if tiempo_bloqueo.total_seconds() < MINUTOS_BLOQUEO * 60:
                minutos_restantes = int((MINUTOS_BLOQUEO * 60 - tiempo_bloqueo.total_seconds()) / 60)
                return True, f"Cuenta bloqueada. Intente nuevamente en {minutos_restantes} minutos.", minutos_restantes
            else:
                # ✅ ACTUALIZADO: Desbloquear automáticamente y registrar en auditoría
                cuenta_bloqueada.bloqueada = False
                cuenta_bloqueada.fecha_desbloqueo = timezone.now()
                cuenta_bloqueada.save()
                
                # Registrar desbloqueo automático en auditoría
                LogAuditoria.objects.create(
                    usuario=user,
                    accion='ACCOUNT_UNLOCKED',
                    tabla_afectada='CuentaBloqueada',
                    registro_id=cuenta_bloqueada.id,
                    ip_address='SYSTEM',  # Sistema automático
                    detalles=f'Cuenta de {user.username} desbloqueada automáticamente después de 30 minutos'
                )
                
                return False, "", 0
        
        return False, "", 0
    except User.DoesNotExist:
        return False, "", 0

# ==========================================
# Nueva función: Registrar intento de login
# ==========================================
def registrar_intento_login(username, ip_address, exitoso, detalles=""):
    """Registra todos los intentos de login en la base de datos"""
    IntentoLogin.objects.create(
        username=username,
        ip_address=ip_address,
        exitoso=exitoso,
        detalles=detalles
    )


# ==========================================
# Nueva función: Verificar intentos fallidos y bloquear cuenta si es necesario
# ==========================================
def verificar_intentos_fallidos(username, ip_address):
    """
    Verifica intentos fallidos en los últimos 15 minutos
    Si hay 5 o más intentos fallidos, bloquea la cuenta
    Retorna: (debe_bloquear: bool, intentos: int)
    """
    INTENTOS_MAXIMOS = 5
    VENTANA_TIEMPO = 15  # minutos
    
    tiempo_limite = timezone.now() - timedelta(minutes=VENTANA_TIEMPO)
    
    # Contar intentos fallidos recientes
    intentos_fallidos = IntentoLogin.objects.filter(
        username=username,
        exitoso=False,
        fecha_hora__gte=tiempo_limite
    ).count()
    
    if intentos_fallidos >= INTENTOS_MAXIMOS:
        try:
            user = User.objects.get(username=username)
            
            # Crear o actualizar registro de cuenta bloqueada
            cuenta_bloqueada, created = CuentaBloqueada.objects.get_or_create(
                usuario=user,
                defaults={
                    'intentos_fallidos': intentos_fallidos,
                    'bloqueada': True,
                    'razon': f'Bloqueada automáticamente por {intentos_fallidos} intentos fallidos'
                }
            )
            
            if not created:
                cuenta_bloqueada.bloqueada = True
                cuenta_bloqueada.intentos_fallidos = intentos_fallidos
                cuenta_bloqueada.fecha_bloqueo = timezone.now()
                cuenta_bloqueada.save()
            
            # Registrar en auditoría
            LogAuditoria.objects.create(
                usuario=user,
                accion='ACCOUNT_LOCKED',
                tabla_afectada='CuentaBloqueada',
                registro_id=cuenta_bloqueada.id,
                ip_address=ip_address,
                detalles=f'Cuenta bloqueada por {intentos_fallidos} intentos fallidos'
            )
            
            return True, intentos_fallidos
        except User.DoesNotExist:
            pass
    
    return False, intentos_fallidos


# ==========================================
# Vista de Login con auditoría y bloqueo de cuentas actualizada
# ==========================================
def login_view(request):
    """Vista de login con auditoría completa y sistema de bloqueo"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        ip_address = obtener_ip_cliente(request)
        
        # 1. Verificar si la cuenta está bloqueada
        bloqueada, mensaje_bloqueo, minutos = verificar_cuenta_bloqueada(username)
        if bloqueada:
            messages.error(request, mensaje_bloqueo)
            registrar_intento_login(username, ip_address, False, "Intento en cuenta bloqueada")
            return render(request, 'registration/login.html')
        
        # 2. Intentar autenticar
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Login exitoso
            login(request, user)
            
            # Registrar intento exitoso
            registrar_intento_login(username, ip_address, True, "Login exitoso")
            
            # Registrar en auditoría
            LogAuditoria.objects.create(
                usuario=user,
                accion='LOGIN',
                tabla_afectada='User',
                registro_id=user.id,
                ip_address=ip_address,
                detalles=f'Login exitoso desde {ip_address}'
            )
            
            # Limpiar bloqueo si existía
            CuentaBloqueada.objects.filter(usuario=user).update(
                bloqueada=False,
                intentos_fallidos=0,
                fecha_desbloqueo=timezone.now()
            )
            
            messages.success(request, f'¡Bienvenido {user.first_name or user.username}!')
            return redirect('dashboard')
        else:
            # Login fallido
            registrar_intento_login(username, ip_address, False, "Credenciales incorrectas")
            
            # Registrar en auditoría (sin usuario porque falló)
            LogAuditoria.objects.create(
                usuario=None,
                accion='LOGIN_FAILED',
                tabla_afectada='User',
                ip_address=ip_address,
                detalles=f'Intento fallido para usuario: {username}'
            )
            
            # Verificar si debe bloquearse la cuenta
            debe_bloquear, intentos = verificar_intentos_fallidos(username, ip_address)
            
            if debe_bloquear:
                messages.error(
                    request, 
                    f'Cuenta bloqueada por múltiples intentos fallidos. '
                    f'Intente nuevamente en 30 minutos.'
                )
            else:
                intentos_restantes = 5 - intentos
                if intentos_restantes <= 2:
                    messages.warning(
                        request, 
                        f'Credenciales incorrectas. Le quedan {intentos_restantes} intentos '
                        f'antes de que su cuenta sea bloqueada.'
                    )
                else:
                    messages.error(request, 'Credenciales incorrectas. Intente nuevamente.')
    
    return render(request, 'registration/login.html')


# ==========================================
# Vista de logout con auditoría actualizada
# ==========================================
@login_required
def logout_view(request):
    """Cierra sesión y registra en auditoría"""
    ip_address = obtener_ip_cliente(request)
    
    LogAuditoria.objects.create(
        usuario=request.user,
        accion='LOGOUT',
        tabla_afectada='User',
        registro_id=request.user.id,
        ip_address=ip_address,
        detalles=f'Logout desde {ip_address}'
    )
    
    logout(request)
    messages.info(request, 'Sesión cerrada correctamente.')
    return redirect('login')


# ==========================================
# VISTAS PRINCIPALES
# ==========================================
@login_required
@requiere_permiso('consultar')
def dashboard(request):
    """Dashboard principal con estadísticas del sistema"""
    from datetime import datetime
    
    # Mensaje de bienvenida según hora del día
    hora_actual = datetime.now().hour
    if hora_actual < 12:
        saludo = "Buenos días"
        icono_saludo = "☀️"
    elif hora_actual < 19:
        saludo = "Buenas tardes"
        icono_saludo = "🌤️"
    else:
        saludo = "Buenas noches"
        icono_saludo = "🌙"
    
    # Estadísticas generales
    total_calificaciones = CalificacionTributaria.objects.filter(activo=True).count()
    total_instrumentos = InstrumentoFinanciero.objects.filter(activo=True).count()
    total_usuarios = User.objects.filter(is_active=True).count()
    
    # Calificaciones por método
    calificaciones_por_metodo = list(
        CalificacionTributaria.objects.filter(activo=True)
        .values('metodo_ingreso')
        .annotate(total=Count('id'))
    )
    
    # Instrumentos por tipo
    instrumentos_por_tipo = list(
        InstrumentoFinanciero.objects.filter(activo=True)
        .values('tipo_instrumento')
        .annotate(total=Count('id'))
    )
    
    # Actividad reciente (últimos 30 días)
    fecha_limite = datetime.now().date() - timedelta(days=30)
    calificaciones_recientes_30d = CalificacionTributaria.objects.filter(
        activo=True,
        fecha_creacion__gte=fecha_limite
    ).count()
    
    # ✅ CORRECCIÓN: Usar el mismo nombre que en el template
    calificaciones_recientes = CalificacionTributaria.objects.filter(
        activo=True
    ).select_related('instrumento', 'usuario_creador').order_by('-fecha_creacion')[:5]
    
    # Logs de auditoría recientes (solo para admin y auditor)
    logs_recientes = None
    try:
        if request.user.is_superuser:
            logs_recientes = LogAuditoria.objects.all().order_by('-fecha_hora')[:5]
        elif hasattr(request.user, 'perfilusuario') and request.user.perfilusuario.rol:
            if request.user.perfilusuario.rol.nombre_rol in ['Administrador', 'Auditor']:
                logs_recientes = LogAuditoria.objects.all().order_by('-fecha_hora')[:5]
    except:
        logs_recientes = None
    
    # Top 5 instrumentos más utilizados
    top_instrumentos = list(
        CalificacionTributaria.objects.filter(activo=True)
        .values(
            'instrumento__codigo_instrumento',
            'instrumento__nombre_instrumento'
        )
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )
    
    # Estadísticas de cargas masivas
    cargas_exitosas = CargaMasiva.objects.filter(estado='EXITOSO').count()
    cargas_fallidas = CargaMasiva.objects.filter(estado='FALLIDO').count()
    
    context = {
        'total_calificaciones': total_calificaciones,
        'total_instrumentos': total_instrumentos,
        'total_usuarios': total_usuarios,
        'calificaciones_recientes_30d': calificaciones_recientes_30d,
        'calificaciones_por_metodo': calificaciones_por_metodo,
        'instrumentos_por_tipo': instrumentos_por_tipo,
        'calificaciones_recientes': calificaciones_recientes,  # ✅ Nombre correcto
        'logs_recientes': logs_recientes,
        'top_instrumentos': top_instrumentos,
        'today': datetime.now(),
        'saludo': saludo,
        'icono_saludo': icono_saludo,
        'cargas_exitosas': cargas_exitosas,
        'cargas_fallidas': cargas_fallidas,
    }
    
    return render(request, 'calificaciones/dashboard.html', context)


# ==========================================
# VISTAS DE CALIFICACIONES
# ==========================================
@login_required
@requiere_permiso('consultar')
def listar_calificaciones(request):
    """Lista todas las calificaciones con filtros y búsqueda"""
    calificaciones = CalificacionTributaria.objects.filter(
        activo=True
    ).select_related('instrumento', 'usuario_creador')
    
    # Filtros
    codigo_instrumento = request.GET.get('codigo_instrumento')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    numero_dj = request.GET.get('numero_dj')
    
    if codigo_instrumento:
        calificaciones = calificaciones.filter(
            instrumento__codigo_instrumento__icontains=codigo_instrumento
        )
    
    if fecha_desde:
        calificaciones = calificaciones.filter(fecha_informe__gte=fecha_desde)
    
    if fecha_hasta:
        calificaciones = calificaciones.filter(fecha_informe__lte=fecha_hasta)
    
    if numero_dj:
        calificaciones = calificaciones.filter(numero_dj__icontains=numero_dj)
    
    calificaciones = calificaciones.order_by('-fecha_creacion')
    
    context = {
        'calificaciones': calificaciones,
        'codigo_instrumento': codigo_instrumento,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'numero_dj': numero_dj,
    }
    
    return render(request, 'calificaciones/listar_calificaciones.html', context)


@login_required
@requiere_permiso('crear')
def crear_calificacion(request):
    """Crea una nueva calificación tributaria"""
    if request.method == 'POST':
        form = CalificacionTributariaForm(request.POST)
        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.usuario_creador = request.user
            calificacion.save()
            
            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion='CREATE',
                tabla_afectada='CalificacionTributaria',
                registro_id=calificacion.id,
                ip_address=ip_address,
                detalles=f'Calificación creada: {calificacion.instrumento.codigo_instrumento}'
            )
            
            messages.success(request, 'Calificación creada exitosamente.')
            return redirect('listar_calificaciones')
    else:
        form = CalificacionTributariaForm()
    
    return render(request, 'calificaciones/form_calificacion.html', {'form': form})


@login_required
@requiere_permiso('modificar')
def editar_calificacion(request, pk):
    """Edita una calificación existente"""
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, activo=True)
    
    if request.method == 'POST':
        form = CalificacionTributariaForm(request.POST, instance=calificacion)
        if form.is_valid():
            calificacion = form.save()
            
            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion='UPDATE',
                tabla_afectada='CalificacionTributaria',
                registro_id=calificacion.id,
                ip_address=ip_address,
                detalles=f'Calificación editada: {calificacion.instrumento.codigo_instrumento}'
            )
            
            messages.success(request, 'Calificación actualizada exitosamente.')
            return redirect('listar_calificaciones')
    else:
        form = CalificacionTributariaForm(instance=calificacion)
    
    context = {'form': form, 'calificacion': calificacion}
    return render(request, 'calificaciones/form_calificacion.html', context)


@login_required
@requiere_permiso('eliminar')
def eliminar_calificacion(request, pk):
    """Eliminación lógica de una calificación"""
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, activo=True)
    
    if request.method == 'POST':
        calificacion.activo = False
        calificacion.save()
        
        # Registrar en auditoría
        ip_address = obtener_ip_cliente(request)
        LogAuditoria.objects.create(
            usuario=request.user,
            accion='DELETE',
            tabla_afectada='CalificacionTributaria',
            registro_id=calificacion.id,
            ip_address=ip_address,
            detalles=f'Calificación eliminada: {calificacion.instrumento.codigo_instrumento}'
        )
        
        messages.success(request, 'Calificación eliminada exitosamente.')
        return redirect('listar_calificaciones')
    
    return render(request, 'calificaciones/confirmar_eliminar.html', {'objeto': calificacion})


# ==========================================
# VISTAS DE INSTRUMENTOS
# ==========================================
@login_required
@requiere_permiso('consultar')
def listar_instrumentos(request):
    """Lista todos los instrumentos financieros"""
    instrumentos = InstrumentoFinanciero.objects.filter(activo=True)
    
    # Búsqueda
    busqueda = request.GET.get('busqueda')
    if busqueda:
        instrumentos = instrumentos.filter(
            Q(codigo_instrumento__icontains=busqueda) |
            Q(nombre_instrumento__icontains=busqueda) |
            Q(tipo_instrumento__icontains=busqueda)
        )
    
    instrumentos = instrumentos.order_by('codigo_instrumento')
    
    return render(request, 'calificaciones/listar_instrumentos.html', {'instrumentos': instrumentos})


@login_required
@requiere_permiso('crear')
def crear_instrumento(request):
    """Crea un nuevo instrumento financiero"""
    if request.method == 'POST':
        form = InstrumentoFinancieroForm(request.POST)
        if form.is_valid():
            instrumento = form.save()
            
            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion='CREATE',
                tabla_afectada='InstrumentoFinanciero',
                registro_id=instrumento.id,
                ip_address=ip_address,
                detalles=f'Instrumento creado: {instrumento.codigo_instrumento}'
            )
            
            messages.success(request, 'Instrumento creado exitosamente.')
            return redirect('listar_instrumentos')
    else:
        form = InstrumentoFinancieroForm()
    
    return render(request, 'calificaciones/form_instrumento.html', {'form': form})


@login_required
@requiere_permiso('modificar')
def editar_instrumento(request, pk):
    """Edita un instrumento existente"""
    instrumento = get_object_or_404(InstrumentoFinanciero, pk=pk, activo=True)
    
    if request.method == 'POST':
        form = InstrumentoFinancieroForm(request.POST, instance=instrumento)
        if form.is_valid():
            instrumento = form.save()
            
            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=request.user,
                accion='UPDATE',
                tabla_afectada='InstrumentoFinanciero',
                registro_id=instrumento.id,
                ip_address=ip_address,
                detalles=f'Instrumento editado: {instrumento.codigo_instrumento}'
            )
            
            messages.success(request, 'Instrumento actualizado exitosamente.')
            return redirect('listar_instrumentos')
    else:
        form = InstrumentoFinancieroForm(instance=instrumento)
    
    context = {'form': form, 'instrumento': instrumento}
    return render(request, 'calificaciones/form_instrumento.html', context)


@login_required
@requiere_permiso('eliminar')
def eliminar_instrumento(request, pk):
    """Eliminación lógica de un instrumento"""
    instrumento = get_object_or_404(InstrumentoFinanciero, pk=pk, activo=True)
    
    # Verificar si tiene calificaciones asociadas
    tiene_calificaciones = CalificacionTributaria.objects.filter(
        instrumento=instrumento, 
        activo=True
    ).exists()
    
    if tiene_calificaciones:
        messages.error(
            request, 
            'No se puede eliminar el instrumento porque tiene calificaciones asociadas.'
        )
        return redirect('listar_instrumentos')
    
    if request.method == 'POST':
        instrumento.activo = False
        instrumento.save()
        
        # Registrar en auditoría
        ip_address = obtener_ip_cliente(request)
        LogAuditoria.objects.create(
            usuario=request.user,
            accion='DELETE',
            tabla_afectada='InstrumentoFinanciero',
            registro_id=instrumento.id,
            ip_address=ip_address,
            detalles=f'Instrumento eliminado: {instrumento.codigo_instrumento}'
        )
        
        messages.success(request, 'Instrumento eliminado exitosamente.')
        return redirect('listar_instrumentos')
    
    return render(request, 'calificaciones/confirmar_eliminar.html', {'objeto': instrumento})


# ==========================================
# CARGA MASIVA
# ==========================================
@login_required
@requiere_permiso('crear')
def carga_masiva(request):
    """Procesa carga masiva de calificaciones desde CSV/Excel"""
    if request.method == 'POST':
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            
            # Crear registro de carga
            carga = CargaMasiva.objects.create(
                usuario=request.user,
                archivo_nombre=archivo.name,
                archivo=archivo,
                estado='PROCESANDO'
            )
            
            try:
                # Detectar tipo de archivo
                if archivo.name.endswith('.xlsx'):
                    registros = procesar_excel(archivo)
                elif archivo.name.endswith('.csv'):
                    registros = procesar_csv(archivo)
                else:
                    raise ValueError('Formato de archivo no soportado')
                
                # Procesar registros
                exitosos = 0
                fallidos = 0
                errores = []
                
                for i, registro in enumerate(registros, start=1):
                    try:
                        # Buscar o crear instrumento
                        instrumento, created = InstrumentoFinanciero.objects.get_or_create(
                            codigo_instrumento=registro['codigo_instrumento'],
                            defaults={
                                'nombre_instrumento': registro.get('nombre_instrumento', ''),
                                'tipo_instrumento': registro.get('tipo_instrumento', 'Otro')
                            }
                        )
                        
                        # Crear calificación
                        CalificacionTributaria.objects.create(
                            instrumento=instrumento,
                            usuario_creador=request.user,
                            monto=Decimal(str(registro['monto'])) if registro.get('monto') else None,
                            factor=Decimal(str(registro['factor'])) if registro.get('factor') else None,
                            metodo_ingreso=registro.get('metodo_ingreso', 'MONTO'),
                            numero_dj=registro.get('numero_dj', ''),
                            fecha_informe=registro['fecha_informe'],
                            observaciones=registro.get('observaciones', '')
                        )
                        
                        exitosos += 1
                    except Exception as e:
                        fallidos += 1
                        errores.append(f'Fila {i}: {str(e)}')
                
                # Actualizar registro de carga
                carga.registros_procesados = len(registros)
                carga.registros_exitosos = exitosos
                carga.registros_fallidos = fallidos
                carga.errores_detalle = '\n'.join(errores)
                
                if fallidos == 0:
                    carga.estado = 'EXITOSO'
                elif exitosos > 0:
                    carga.estado = 'PARCIAL'
                else:
                    carga.estado = 'FALLIDO'
                
                carga.save()
                
                # Registrar en auditoría
                ip_address = obtener_ip_cliente(request)
                LogAuditoria.objects.create(
                    usuario=request.user,
                    accion='CREATE',
                    tabla_afectada='CargaMasiva',
                    registro_id=carga.id,
                    ip_address=ip_address,
                    detalles=f'Carga masiva: {exitosos} exitosos, {fallidos} fallidos'
                )
                
                messages.success(
                    request,
                    f'Procesados {exitosos} registros correctamente. {fallidos} con errores.'
                )
                
            except Exception as e:
                carga.estado = 'FALLIDO'
                carga.errores_detalle = str(e)
                carga.save()
                messages.error(request, f'Error al procesar archivo: {str(e)}')
            
            return redirect('dashboard')
    else:
        form = CargaMasivaForm()
    
    return render(request, 'calificaciones/carga_masiva.html', {'form': form})


def procesar_excel(archivo):
    """Procesa archivo Excel y retorna lista de registros"""
    wb = openpyxl.load_workbook(archivo)
    sheet = wb.active
    registros = []
    
    headers = [cell.value for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        registro = dict(zip(headers, row))
        if registro.get('codigo_instrumento'):
            registros.append(registro)
    
    return registros


def procesar_csv(archivo):
    """Procesa archivo CSV y retorna lista de registros"""
    contenido = archivo.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(contenido))
    return list(reader)


# ==========================================
# EXPORTACIÓN DE DATOS
# ==========================================
@login_required
@requiere_permiso('consultar')
def exportar_excel(request):
    """Exporta calificaciones a Excel"""
    calificaciones = CalificacionTributaria.objects.filter(
        activo=True
    ).select_related('instrumento', 'usuario_creador')
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calificaciones"
    
    # Encabezados
    headers = [
        'ID', 'Código Instrumento', 'Nombre Instrumento', 'Monto', 
        'Factor', 'Método Ingreso', 'Número DJ', 'Fecha Informe', 
        'Usuario Creador', 'Fecha Creación', 'Observaciones'
    ]
    ws.append(headers)
    
    # Datos
    for cal in calificaciones:
        ws.append([
            cal.id,
            cal.instrumento.codigo_instrumento,
            cal.instrumento.nombre_instrumento,
            float(cal.monto) if cal.monto else None,
            float(cal.factor) if cal.factor else None,
            cal.metodo_ingreso,
            cal.numero_dj,
            cal.fecha_informe.strftime('%Y-%m-%d'),
            cal.usuario_creador.username if cal.usuario_creador else '',
            cal.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
            cal.observaciones
        ])
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=calificaciones_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    
    # Registrar en auditoría
    ip_address = obtener_ip_cliente(request)
    LogAuditoria.objects.create(
        usuario=request.user,
        accion='READ',
        tabla_afectada='CalificacionTributaria',
        ip_address=ip_address,
        detalles='Exportación de calificaciones a Excel'
    )
    
    return response


@login_required
@requiere_permiso('consultar')
def exportar_csv(request):
    """Exporta calificaciones a CSV"""
    calificaciones = CalificacionTributaria.objects.filter(
        activo=True
    ).select_related('instrumento', 'usuario_creador')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=calificaciones_{timezone.now().strftime("%Y%m%d")}.csv'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Código Instrumento', 'Nombre Instrumento', 'Monto', 
        'Factor', 'Método Ingreso', 'Número DJ', 'Fecha Informe', 
        'Usuario Creador', 'Fecha Creación', 'Observaciones'
    ])
    
    for cal in calificaciones:
        writer.writerow([
            cal.id,
            cal.instrumento.codigo_instrumento,
            cal.instrumento.nombre_instrumento,
            cal.monto,
            cal.factor,
            cal.metodo_ingreso,
            cal.numero_dj,
            cal.fecha_informe,
            cal.usuario_creador.username if cal.usuario_creador else '',
            cal.fecha_creacion,
            cal.observaciones
        ])
    
    # Registrar en auditoría
    ip_address = obtener_ip_cliente(request)
    LogAuditoria.objects.create(
        usuario=request.user,
        accion='READ',
        tabla_afectada='CalificacionTributaria',
        ip_address=ip_address,
        detalles='Exportación de calificaciones a CSV'
    )
    
    return response


# ==========================================
# PERFIL DE USUARIO
# ==========================================
@login_required
def mi_perfil(request):
    """Muestra y permite editar el perfil del usuario actual"""
    try:
        perfil = request.user.perfilusuario
    except PerfilUsuario.DoesNotExist:
        # Crear perfil si no existe
        perfil = PerfilUsuario.objects.create(usuario=request.user)
    
    if request.method == 'POST':
        # Actualizar información básica
        request.user.first_name = request.POST.get('first_name', '')
        request.user.last_name = request.POST.get('last_name', '')
        request.user.email = request.POST.get('email', '')
        request.user.save()
        
        # Actualizar perfil
        perfil.telefono = request.POST.get('telefono', '')
        perfil.departamento = request.POST.get('departamento', '')
        perfil.save()
        
        messages.success(request, 'Perfil actualizado exitosamente.')
        return redirect('mi_perfil')
    
    context = {
        'perfil': perfil,
    }
    
    return render(request, 'calificaciones/mi_perfil.html', context)


# ==========================================
# REGISTRO DE USUARIOS
# ==========================================
def registro(request):
    """Permite el registro de nuevos usuarios"""
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Crear perfil de usuario con rol por defecto (Auditor)
            rol_auditor = Rol.objects.get(nombre_rol='Auditor')
            PerfilUsuario.objects.create(
                usuario=user,
                rol=rol_auditor
            )
            
            # Registrar en auditoría
            ip_address = obtener_ip_cliente(request)
            LogAuditoria.objects.create(
                usuario=user,
                accion='CREATE',
                tabla_afectada='User',
                registro_id=user.id,
                ip_address=ip_address,
                detalles=f'Usuario registrado: {user.username}'
            )
            
            messages.success(request, 'Registro exitoso. Ya puedes iniciar sesión.')
            return redirect('login')
    else:
        form = RegistroForm()
    
    return render(request, 'calificaciones/registro.html', {'form': form})


# ==========================================
# AUDITORÍA
# ==========================================
@login_required
@requiere_permiso('consultar')
def registro_auditoria(request):
    """Muestra el registro completo de auditoría (solo para Administrador y Auditor)"""
    logs = LogAuditoria.objects.all().select_related('usuario').order_by('-fecha_hora')
    
    # Filtros
    usuario_id = request.GET.get('usuario')
    accion = request.GET.get('accion')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if usuario_id:
        logs = logs.filter(usuario_id=usuario_id)
    
    if accion:
        logs = logs.filter(accion=accion)
    
    if fecha_desde:
        logs = logs.filter(fecha_hora__date__gte=fecha_desde)
    
    if fecha_hasta:
        logs = logs.filter(fecha_hora__date__lte=fecha_hasta)
    
    # Limitar a últimos 1000 registros para performance
    logs = logs[:1000]
    
    # Lista de usuarios para filtro
    usuarios = User.objects.filter(is_active=True).order_by('username')
    
    # ✅ Lista de acciones disponibles
    acciones = LogAuditoria.ACCIONES
    
    context = {
        'logs': logs,
        'usuarios': usuarios,
        'acciones': acciones,
        'usuario_filtro': usuario_id,
        'accion_filtro': accion,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'calificaciones/registro_auditoria.html', context)
