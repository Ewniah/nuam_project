"""
Script para generar archivo Excel con datos de prueba para carga masiva.

Genera un archivo Excel con calificaciones tributarias de ejemplo que puede
ser utilizado para probar la funcionalidad de carga masiva del sistema NUAM.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random
from decimal import Decimal


def generar_excel_prueba():
    """
    Genera archivo Excel con datos de prueba para carga masiva.
    
    Crea un workbook con 20 registros de ejemplo que incluyen:
    - 5 tipos diferentes de instrumentos financieros
    - Variedad en montos y factores
    - Fechas distribuidas en los últimos 60 días
    - Algunos registros con observaciones
    """
    # Crear workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Calificaciones"

    # Headers con formato
    headers = [
        "codigo_instrumento",
        "nombre_instrumento", 
        "tipo_instrumento",
        "monto",
        "factor",
        "metodo_ingreso",
        "numero_dj",
        "fecha_informe",
        "observaciones"
    ]

    # Estilo de headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos de ejemplo
    instrumentos_base = [
        {
            "codigo": "BCH-2024-001",
            "nombre": "Banco de Chile Bono Subordinado",
            "tipo": "Bono"
        },
        {
            "codigo": "SANTANDER-2024-002",
            "nombre": "Banco Santander Depósito a Plazo",
            "tipo": "Depósito"
        },
        {
            "codigo": "BCI-2024-003",
            "nombre": "BCI Línea de Crédito Empresarial",
            "tipo": "Crédito"
        },
        {
            "codigo": "ITAU-2024-004",
            "nombre": "Itaú Pagaré Financiero",
            "tipo": "Pagaré"
        },
        {
            "codigo": "SECURITY-2024-005",
            "nombre": "Security Letra Hipotecaria",
            "tipo": "Letra"
        }
    ]

    # Generar 20 registros
    fecha_base = datetime.now()
    
    registros = []
    for i in range(20):
        instrumento = instrumentos_base[i % len(instrumentos_base)]
        
        # Variar fechas en los últimos 60 días
        dias_atras = random.randint(0, 60)
        fecha_informe = (fecha_base - timedelta(days=dias_atras)).strftime("%Y-%m-%d")
        
        # Alternar entre MONTO y FACTOR
        metodo = "MONTO" if i % 3 != 0 else "FACTOR"
        
        if metodo == "MONTO":
            monto = round(random.uniform(1000000, 50000000), 2)
            factor = None
        else:
            monto = None
            factor = round(random.uniform(0.01, 0.50), 4)
        
        # Generar número DJ único (max 10 caracteres)
        numero_dj = f"DJ{i+1:04d}"  # Formato: DJ0001, DJ0002, etc.
        
        # Algunas observaciones
        observaciones = ""
        if i % 5 == 0:
            obs_opciones = [
                "Instrumento de alta liquidez",
                "Requiere revisión trimestral",
                "Aprobado por comité de riesgos",
                "Pendiente de ratificación",
                "Calificación preliminar"
            ]
            observaciones = obs_opciones[i % len(obs_opciones)]
        
        registro = {
            "codigo_instrumento": f"{instrumento['codigo']}-{i+1:03d}",
            "nombre_instrumento": instrumento["nombre"],
            "tipo_instrumento": instrumento["tipo"],
            "monto": monto,
            "factor": factor,
            "metodo_ingreso": metodo,
            "numero_dj": numero_dj,
            "fecha_informe": fecha_informe,
            "observaciones": observaciones
        }
        registros.append(registro)

    # Escribir datos
    for row_num, registro in enumerate(registros, 2):
        sheet.cell(row=row_num, column=1).value = registro["codigo_instrumento"]
        sheet.cell(row=row_num, column=2).value = registro["nombre_instrumento"]
        sheet.cell(row=row_num, column=3).value = registro["tipo_instrumento"]
        sheet.cell(row=row_num, column=4).value = registro["monto"]
        sheet.cell(row=row_num, column=5).value = registro["factor"]
        sheet.cell(row=row_num, column=6).value = registro["metodo_ingreso"]
        sheet.cell(row=row_num, column=7).value = registro["numero_dj"]
        sheet.cell(row=row_num, column=8).value = registro["fecha_informe"]
        sheet.cell(row=row_num, column=9).value = registro["observaciones"]

    # Ajustar anchos de columna
    column_widths = {
        'A': 25,  # codigo_instrumento
        'B': 40,  # nombre_instrumento
        'C': 15,  # tipo_instrumento
        'D': 15,  # monto
        'E': 12,  # factor
        'F': 18,  # metodo_ingreso
        'G': 18,  # numero_dj
        'H': 15,  # fecha_informe
        'I': 40   # observaciones
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Guardar archivo
    filename = "datos_prueba_carga_masiva.xlsx"
    wb.save(filename)
    
    print(f"✅ Archivo '{filename}' generado exitosamente")
    print(f"📊 Total de registros: {len(registros)}")
    print(f"📁 Ubicación: {filename}")
    print("\nResumen de datos generados:")
    print(f"  - Registros con MONTO: {sum(1 for r in registros if r['monto'] is not None)}")
    print(f"  - Registros con FACTOR: {sum(1 for r in registros if r['factor'] is not None)}")
    print(f"  - Registros con observaciones: {sum(1 for r in registros if r['observaciones'])}")
    print(f"  - Tipos de instrumentos: {len(instrumentos_base)}")
    print("\n🚀 Archivo listo para carga masiva en el sistema")


if __name__ == "__main__":
    generar_excel_prueba()
