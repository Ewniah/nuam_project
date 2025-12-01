"""
Script de Prueba de Estrés - 30 Factores Tributarios (8-37)
===========================================================

Propósito:
    Generar archivo Excel con 30 factores individuales para verificar:
    1. Validación de rango individual (0 <= factor <= 1)
    2. Validación de suma de factores 8-16 (suma <= 1.0)
    3. Mapeo dinámico de 30 columnas en carga masiva

Autor: QA Automation
Fecha: 2025-12-01
Versión: 1.0 - Stress Test 30 Factores
"""

import os
import sys
from datetime import date
from decimal import Decimal

# Agregar directorio raíz al path para imports de Django
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nuam_project.settings')
import django
django.setup()

# Imports después de Django setup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from calificaciones.models import InstrumentoFinanciero

# ==============================================================================
# CONFIGURACIÓN
# ==============================================================================

OUTPUT_FILE = "test_30factores_stress.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), OUTPUT_FILE)

# Rango de factores a probar (8-37 = 30 factores)
FACTOR_RANGE = range(8, 38)

# ==============================================================================
# FUNCIONES AUXILIARES
# ==============================================================================

def crear_instrumento_prueba(codigo, nombre, tipo):
    """
    Crea o recupera un instrumento financiero para las pruebas.
    
    Args:
        codigo (str): Código del instrumento
        nombre (str): Nombre descriptivo
        tipo (str): Tipo de instrumento
    
    Returns:
        InstrumentoFinanciero: Instancia del instrumento
    """
    instrumento, created = InstrumentoFinanciero.objects.get_or_create(
        codigo_instrumento=codigo,
        defaults={
            'nombre_instrumento': nombre,
            'tipo_instrumento': tipo,
            'activo': True
        }
    )
    if created:
        print(f"✓ Instrumento '{codigo}' creado")
    else:
        print(f"ℹ Instrumento '{codigo}' ya existe")
    
    return instrumento


def aplicar_formato_excel(ws):
    """
    Aplica formato profesional a la hoja Excel.
    
    Args:
        ws: Worksheet de openpyxl
    """
    # Estilo de encabezado
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formato a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Ajustar anchos de columna usando get_column_letter para soportar >26 columnas
    ws.column_dimensions['A'].width = 20  # codigo_instrumento
    ws.column_dimensions['B'].width = 15  # metodo_ingreso
    ws.column_dimensions['C'].width = 12  # numero_dj
    ws.column_dimensions['D'].width = 15  # fecha_informe
    
    # Metadata (columnas E-J: 6 campos)
    for col_num in range(5, 11):  # E(5) a J(10)
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 14
    
    # Factores 8-37 (columnas K en adelante: 30 factores)
    for i, factor_num in enumerate(FACTOR_RANGE, start=11):  # Empieza en K(11)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 11
    
    # Observaciones (última columna: después de 4 base + 6 metadata + 30 factores = col 41)
    last_col_letter = get_column_letter(41)
    ws.column_dimensions[last_col_letter].width = 40


# ==============================================================================
# ESCENARIOS DE PRUEBA
# ==============================================================================

def generar_escenario_1_golden():
    """
    ESCENARIO 1: GOLDEN RECORD (30 Factores Válidos)
    =================================================
    - Todos los factores = 0.01 (válido: 0 < 0.01 < 1)
    - Suma de factores 8-16 = 0.09 (válido: < 1.0)
    - Debe insertarse exitosamente ✓
    
    Returns:
        dict: Datos del registro
    """
    print(f"\n{'='*70}")
    print(f"ESCENARIO 1: GOLDEN RECORD - 30 FACTORES VÁLIDOS")
    print(f"{'='*70}")
    
    # Crear diccionario base
    registro = {
        'codigo_instrumento': 'FULL-OK-001',
        'metodo_ingreso': 'FACTOR',
        'numero_dj': 'DJ1949',
        'fecha_informe': '2025-12-01',
        # Metadata (TASK-004 - HDU_Inacap.xlsx)
        'secuencia': 1,
        'numero_dividendo': 0,
        'tipo_sociedad': 'A',
        'valor_historico': 1000000.50,
        'mercado': 'ACN',
        'ejercicio': 2025,
    }
    
    # Todos los factores = 0.01
    for factor_num in FACTOR_RANGE:
        registro[f'factor_{factor_num}'] = 0.01
    
    # Calcular suma de factores críticos (8-16)
    suma_criticos = sum(registro[f'factor_{i}'] for i in range(8, 17))
    
    print(f"Total de factores: {len(FACTOR_RANGE)}")
    print(f"Valor de cada factor: 0.01")
    print(f"Suma de factores 8-16: {suma_criticos:.2f} (debe ser <= 1.0)")
    print(f"Estado: {'✓ VÁLIDO' if suma_criticos <= 1.0 else '✗ INVÁLIDO'}")
    
    registro['observaciones'] = f'Escenario 1: Golden Record - Todos los {len(FACTOR_RANGE)} factores válidos'
    
    return registro


def generar_escenario_2_range_failure():
    """
    ESCENARIO 2: FALLO DE VALIDACIÓN DE RANGO
    ==========================================
    - factor_37 = 5.0 (INVÁLIDO: > 1)
    - Resto de factores = 0.01 (válidos)
    - Debe fallar con ValidationError ✗
    - Error esperado: "El factor_37 debe estar entre 0 y 1. Valor recibido: 5.0"
    
    Returns:
        dict: Datos del registro inválido
    """
    print(f"\n{'='*70}")
    print(f"ESCENARIO 2: FALLO VALIDACIÓN DE RANGO")
    print(f"{'='*70}")
    
    registro = {
        'codigo_instrumento': 'FULL-FAIL-RANGE',
        'metodo_ingreso': 'FACTOR',
        'numero_dj': 'DJ1922',
        'fecha_informe': '2025-12-01',
        # Metadata (TASK-004 - HDU_Inacap.xlsx)
        'secuencia': 2,
        'numero_dividendo': 1,
        'tipo_sociedad': 'C',
        'valor_historico': 500000.00,
        'mercado': 'BVS',
        'ejercicio': 2025,
    }
    
    # Todos los factores válidos excepto factor_37
    for factor_num in FACTOR_RANGE:
        if factor_num == 37:
            registro[f'factor_{factor_num}'] = 5.0  # INVÁLIDO
        else:
            registro[f'factor_{factor_num}'] = 0.01
    
    print(f"factor_37: 5.0 (INVÁLIDO - Fuera de rango 0-1)")
    print(f"Resto de factores: 0.01 (válidos)")
    print(f"Estado: ✗ INVÁLIDO - Violación de REGLA A (rango individual)")
    print(f"Error esperado: ValidationError")
    
    registro['observaciones'] = 'Escenario 2: Range Failure - factor_37 = 5.0 (debe fallar)'
    
    return registro


def generar_escenario_3_sum_failure():
    """
    ESCENARIO 3: FALLO DE VALIDACIÓN DE SUMA
    =========================================
    - factores 8-16 = 0.2 cada uno (9 factores)
    - Suma = 9 × 0.2 = 1.8 (INVÁLIDO: > 1.0)
    - Resto de factores = 0.01
    - Debe fallar con ValidationError ✗
    - Error esperado: "La suma de los factores 8 al 16 no puede superar 1."
    
    Returns:
        dict: Datos del registro inválido
    """
    print(f"\n{'='*70}")
    print(f"ESCENARIO 3: FALLO VALIDACIÓN DE SUMA")
    print(f"{'='*70}")
    
    registro = {
        'codigo_instrumento': 'FULL-FAIL-SUM',
        'metodo_ingreso': 'FACTOR',
        'numero_dj': 'DJ1949',
        'fecha_informe': '2025-12-01',
        # Metadata (TASK-004 - HDU_Inacap.xlsx)
        'secuencia': 3,
        'numero_dividendo': 2,
        'tipo_sociedad': 'A',
        'valor_historico': 750000.25,
        'mercado': 'ACN',
        'ejercicio': 2025,
    }
    
    # Factores 8-16 = 0.2 (suma excede 1.0)
    # Resto = 0.01
    for factor_num in FACTOR_RANGE:
        if 8 <= factor_num <= 16:
            registro[f'factor_{factor_num}'] = 0.2  # INVÁLIDO en suma
        else:
            registro[f'factor_{factor_num}'] = 0.01
    
    suma_criticos = sum(registro[f'factor_{i}'] for i in range(8, 17))
    
    print(f"Factores 8-16: 0.2 cada uno")
    print(f"Cantidad de factores críticos: 9")
    print(f"Suma de factores 8-16: {suma_criticos:.1f} (debe ser <= 1.0)")
    print(f"Estado: ✗ INVÁLIDO - Violación de REGLA B (suma límite)")
    print(f"Error esperado: ValidationError")
    
    registro['observaciones'] = 'Escenario 3: Sum Failure - Suma de factores 8-16 = 1.8 (debe fallar)'
    
    return registro


# ==============================================================================
# GENERACIÓN DEL ARCHIVO EXCEL
# ==============================================================================

def generar_excel_prueba():
    """
    Genera el archivo Excel con los 3 escenarios de prueba.
    Usa generación dinámica de columnas para los 30 factores.
    """
    print(f"\n{'='*70}")
    print(f"GENERANDO ARCHIVO DE PRUEBA: {OUTPUT_FILE}")
    print(f"{'='*70}")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test 30 Factores"
    
    # GENERACIÓN DINÁMICA DE ENCABEZADOS
    headers = [
        'codigo_instrumento',
        'metodo_ingreso',
        'numero_dj',
        'fecha_informe',
        # Campos metadata (TASK-004 - HDU_Inacap.xlsx Hoja 3.1)
        'secuencia',
        'numero_dividendo',
        'tipo_sociedad',
        'valor_historico',
        'mercado',
        'ejercicio',
    ]
    
    # Agregar dinámicamente factor_8 hasta factor_37
    for factor_num in FACTOR_RANGE:
        headers.append(f'factor_{factor_num}')
    
    # Agregar observaciones al final
    headers.append('observaciones')
    
    print(f"\nColumnas generadas: {len(headers)}")
    print(f"  - Campos base: 4")
    print(f"  - Metadata (TASK-004 HDU): 6")
    print(f"  - Factores (8-37): {len(FACTOR_RANGE)}")
    print(f"  - Observaciones: 1")
    print(f"  - Total: {len(headers)}")
    
    # Escribir encabezados
    ws.append(headers)
    
    # Generar escenarios
    escenarios = [
        generar_escenario_1_golden(),
        generar_escenario_2_range_failure(),
        generar_escenario_3_sum_failure(),
    ]
    
    # Escribir datos de cada escenario
    for escenario in escenarios:
        fila = []
        for header in headers:
            valor = escenario.get(header, '')
            fila.append(valor)
        ws.append(fila)
    
    # Aplicar formato
    aplicar_formato_excel(ws)
    
    # Guardar archivo
    wb.save(OUTPUT_PATH)
    
    print(f"\n{'='*70}")
    print(f"✓ Archivo generado exitosamente: {OUTPUT_PATH}")
    print(f"{'='*70}")


# ==============================================================================
# PREPARACIÓN DEL ENTORNO
# ==============================================================================

def preparar_instrumentos():
    """
    Crea los instrumentos financieros necesarios para las pruebas.
    """
    print(f"\n{'='*70}")
    print(f"PREPARANDO INSTRUMENTOS DE PRUEBA")
    print(f"{'='*70}")
    
    instrumentos = [
        ('FULL-OK-001', 'Instrumento 30 Factores OK', 'Bono'),
        ('FULL-FAIL-RANGE', 'Instrumento Fallo Rango', 'Acción'),
        ('FULL-FAIL-SUM', 'Instrumento Fallo Suma', 'Fondo'),
    ]
    
    for codigo, nombre, tipo in instrumentos:
        crear_instrumento_prueba(codigo, nombre, tipo)


# ==============================================================================
# FUNCIÓN PRINCIPAL
# ==============================================================================

def main():
    """
    Función principal del script.
    """
    print(f"\n{'#'*70}")
    print(f"#  SCRIPT DE PRUEBA DE ESTRÉS - 30 FACTORES TRIBUTARIOS")
    print(f"#  Fecha: {date.today()}")
    print(f"#  Rango de factores: {FACTOR_RANGE.start}-{FACTOR_RANGE.stop - 1}")
    print(f"{'#'*70}")
    
    try:
        # Paso 1: Preparar instrumentos
        preparar_instrumentos()
        
        # Paso 2: Generar archivo Excel
        generar_excel_prueba()
        
        # Resumen final
        print(f"\n{'='*70}")
        print(f"RESUMEN DE PRUEBAS DE ESTRÉS")
        print(f"{'='*70}")
        print(f"✓ Archivo creado: {OUTPUT_FILE}")
        print(f"✓ Total de registros: 3")
        print(f"✓ Total de factores por registro: {len(FACTOR_RANGE)}")
        print(f"✓ Total de columnas: {4 + 6 + len(FACTOR_RANGE) + 1}")  # Base + Metadata(6) + Factores + Obs
        print(f"")
        print(f"RESULTADOS ESPERADOS AL CARGAR EL ARCHIVO:")
        print(f"  - Fila 2 (GOLDEN):       ✓ Éxito (1 registro)")
        print(f"  - Fila 3 (RANGE FAIL):   ✗ Error de rango (factor_37 = 5.0)")
        print(f"  - Fila 4 (SUM FAIL):     ✗ Error de suma (factores 8-16 > 1.0)")
        print(f"")
        print(f"ESTADÍSTICAS ESPERADAS:")
        print(f"  - Procesados: 3")
        print(f"  - Exitosos: 1")
        print(f"  - Fallidos: 2")
        print(f"  - Estado: PARCIAL")
        print(f"")
        print(f"VALIDACIONES A VERIFICAR:")
        print(f"  ✓ REGLA A: Rango individual (0 <= factor <= 1) para 30 factores")
        print(f"  ✓ REGLA B: Suma de factores 8-16 <= 1.0")
        print(f"  ✓ Mapeo dinámico de 30 columnas en carga_masiva()")
        print(f"")
        print(f"PRÓXIMOS PASOS:")
        print(f"  1. Ir a http://127.0.0.1:8000/carga-masiva/")
        print(f"  2. Subir el archivo: {OUTPUT_FILE}")
        print(f"  3. Verificar que solo 1 registro se carga exitosamente")
        print(f"  4. Revisar errores detallados en el historial")
        print(f"  5. Confirmar que los mensajes de error sean específicos")
        print(f"{'='*70}\n")
        
        return 0
        
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit_code = main()
    sys.exit(exit_code)
