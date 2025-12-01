"""
Script de Prueba Final - Validaciones de Negocio y Restricciones de Unicidad
=============================================================================

Propósito:
    Generar archivo Excel de prueba con 3 escenarios para verificar:
    1. Registro válido (Golden Record)
    2. Fallo de validación matemática (suma factores > 1.0)
    3. Detección de duplicados (unique_together constraint)

Autor: QA Automation
Fecha: 2025-12-01
Versión: 1.0
"""

import os
import sys
from datetime import date
from decimal import Decimal

# Agregar directorio raíz al path para imports de Django
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nuam_project.settings')
import django
django.setup()

# Imports de Django después de setup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from calificaciones.models import InstrumentoFinanciero

# ==============================================================================
# CONFIGURACIÓN DE PRUEBAS
# ==============================================================================

OUTPUT_FILE = "test_validaciones_final.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), OUTPUT_FILE)

# ==============================================================================
# FUNCIONES AUXILIARES
# ==============================================================================

def crear_instrumento_prueba(codigo, nombre, tipo):
    """
    Crea o recupera un instrumento financiero para las pruebas.
    
    Args:
        codigo (str): Código del instrumento (ej: TEST001)
        nombre (str): Nombre descriptivo
        tipo (str): Tipo de instrumento
    
    Returns:
        InstrumentoFinanciero: Instancia del instrumento
    """
    instrumento, created = InstrumentoFinanciero.objects.get_or_create(
        codigo_instrumento=codigo,
        defaults={
            'nombre_instrumento': nombre,
            'tipo_instrumento': tipo,
            'activo': True
        }
    )
    if created:
        print(f"✓ Instrumento '{codigo}' creado exitosamente")
    else:
        print(f"ℹ Instrumento '{codigo}' ya existe en el sistema")
    
    return instrumento


def calcular_factores(montos):
    """
    Calcula factores a partir de montos (Factor = Monto / Suma Total).
    
    Args:
        montos (list): Lista de 5 montos [monto_8, monto_9, ..., monto_12]
    
    Returns:
        list: Lista de 5 factores calculados
    """
    suma_total = sum(montos)
    if suma_total == 0:
        return [0, 0, 0, 0, 0]
    
    factores = [Decimal(str(m / suma_total)) for m in montos]
    return factores


def aplicar_formato_excel(ws):
    """
    Aplica formato profesional a la hoja Excel.
    
    Args:
        ws: Worksheet de openpyxl
    """
    # Formato de encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 18,  # codigo_instrumento
        'B': 15,  # metodo_ingreso
        'C': 15,  # monto_8
        'D': 15,  # monto_9
        'E': 15,  # monto_10
        'F': 15,  # monto_11
        'G': 15,  # monto_12
        'H': 15,  # factor_8
        'I': 15,  # factor_9
        'J': 15,  # factor_10
        'K': 15,  # factor_11
        'L': 15,  # factor_12
        'M': 12,  # numero_dj
        'N': 15,  # fecha_informe
        'O': 30,  # observaciones
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


# ==============================================================================
# ESCENARIOS DE PRUEBA
# ==============================================================================

def generar_escenario_1_golden():
    """
    ESCENARIO 1: GOLDEN RECORD (Registro Válido)
    ============================================
    - Suma de factores < 1.0 ✓
    - Todos los datos válidos ✓
    - Debe insertarse exitosamente ✓
    
    Returns:
        dict: Datos del registro
    """
    # Montos que suman menos de 1.0 cuando se convierten a factores
    montos = [100000, 150000, 200000, 120000, 80000]  # Total: 650,000
    factores = calcular_factores(montos)
    suma_factores = sum(factores)
    
    print(f"\n{'='*70}")
    print(f"ESCENARIO 1: GOLDEN RECORD")
    print(f"{'='*70}")
    print(f"Montos: {montos}")
    print(f"Factores calculados: {[f'{f:.8f}' for f in factores]}")
    print(f"Suma de factores: {suma_factores:.8f} (debe ser < 1.0)")
    print(f"Estado: {'✓ VÁLIDO' if suma_factores < 1.0 else '✗ INVÁLIDO'}")
    
    return {
        'codigo_instrumento': 'TEST001',
        'metodo_ingreso': 'MONTO',
        'monto_8': montos[0],
        'monto_9': montos[1],
        'monto_10': montos[2],
        'monto_11': montos[3],
        'monto_12': montos[4],
        'factor_8': float(factores[0]),
        'factor_9': float(factores[1]),
        'factor_10': float(factores[2]),
        'factor_11': float(factores[3]),
        'factor_12': float(factores[4]),
        'numero_dj': 'DJ1949',
        'fecha_informe': '2025-12-01',
        'observaciones': 'Escenario 1: Golden Record - Todos los valores son válidos',
    }


def generar_escenario_2_math_failure():
    """
    ESCENARIO 2: FALLO DE VALIDACIÓN MATEMÁTICA
    ===========================================
    - Suma de factores > 1.0 ✗
    - Debe fallar con ValidationError ✗
    - Error esperado: "La suma de los factores 8 al 16 no puede superar 1."
    
    Returns:
        dict: Datos del registro inválido
    """
    # Factores que intencionalmente suman más de 1.0
    factores = [0.30, 0.25, 0.28, 0.22, 0.15]  # Suma: 1.20 > 1.0
    suma_factores = sum(factores)
    
    print(f"\n{'='*70}")
    print(f"ESCENARIO 2: FALLO VALIDACIÓN MATEMÁTICA")
    print(f"{'='*70}")
    print(f"Factores: {factores}")
    print(f"Suma de factores: {suma_factores:.2f} (debe ser < 1.0)")
    print(f"Estado: {'✗ INVÁLIDO - Suma excede 1.0' if suma_factores > 1.0 else '✓ VÁLIDO'}")
    print(f"Error esperado: ValidationError")
    
    return {
        'codigo_instrumento': 'TEST002',
        'metodo_ingreso': 'FACTOR',
        'monto_8': '',
        'monto_9': '',
        'monto_10': '',
        'monto_11': '',
        'monto_12': '',
        'factor_8': factores[0],
        'factor_9': factores[1],
        'factor_10': factores[2],
        'factor_11': factores[3],
        'factor_12': factores[4],
        'numero_dj': 'DJ1922',
        'fecha_informe': '2025-12-01',
        'observaciones': 'Escenario 2: Math Failure - Suma de factores excede 1.0 (debería fallar)',
    }


def generar_escenario_3_duplicate():
    """
    ESCENARIO 3: DETECCIÓN DE DUPLICADOS
    ====================================
    - Registro idéntico al Escenario 1 en campos unique_together
    - unique_together = ['instrumento', 'fecha_informe', 'numero_dj']
    - Debe fallar con IntegrityError (duplicado) ✗
    - Error esperado: "Este registro ya existe en el sistema (Duplicado)."
    
    Returns:
        dict: Datos del registro duplicado
    """
    # Mismos valores que Escenario 1 en campos únicos
    montos = [100000, 150000, 200000, 120000, 80000]
    factores = calcular_factores(montos)
    
    print(f"\n{'='*70}")
    print(f"ESCENARIO 3: DETECCIÓN DE DUPLICADOS")
    print(f"{'='*70}")
    print(f"Instrumento: TEST001 (igual que Escenario 1)")
    print(f"Fecha Informe: 2025-12-01 (igual que Escenario 1)")
    print(f"Número DJ: DJ1949 (igual que Escenario 1)")
    print(f"Estado: ✗ DUPLICADO - Debe ser rechazado por unique_together")
    print(f"Error esperado: IntegrityError (Duplicate)")
    
    return {
        'codigo_instrumento': 'TEST001',  # DUPLICADO: mismo instrumento
        'metodo_ingreso': 'MONTO',
        'monto_8': montos[0],
        'monto_9': montos[1],
        'monto_10': montos[2],
        'monto_11': montos[3],
        'monto_12': montos[4],
        'factor_8': float(factores[0]),
        'factor_9': float(factores[1]),
        'factor_10': float(factores[2]),
        'factor_11': float(factores[3]),
        'factor_12': float(factores[4]),
        'numero_dj': 'DJ1949',  # DUPLICADO: mismo DJ
        'fecha_informe': '2025-12-01',  # DUPLICADO: misma fecha
        'observaciones': 'Escenario 3: Duplicate - Mismo instrumento/fecha/DJ que Escenario 1 (debería fallar)',
    }


# ==============================================================================
# GENERACIÓN DEL ARCHIVO EXCEL
# ==============================================================================

def generar_excel_prueba():
    """
    Genera el archivo Excel con los 3 escenarios de prueba.
    """
    print(f"\n{'='*70}")
    print(f"GENERANDO ARCHIVO DE PRUEBA: {OUTPUT_FILE}")
    print(f"{'='*70}")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Validaciones"
    
    # Encabezados (según formato de carga masiva)
    headers = [
        'codigo_instrumento',
        'metodo_ingreso',
        'monto_8',
        'monto_9',
        'monto_10',
        'monto_11',
        'monto_12',
        'factor_8',
        'factor_9',
        'factor_10',
        'factor_11',
        'factor_12',
        'numero_dj',
        'fecha_informe',
        'observaciones',
    ]
    ws.append(headers)
    
    # Generar escenarios
    escenarios = [
        generar_escenario_1_golden(),
        generar_escenario_2_math_failure(),
        generar_escenario_3_duplicate(),
    ]
    
    # Escribir datos
    for escenario in escenarios:
        fila = [
            escenario['codigo_instrumento'],
            escenario['metodo_ingreso'],
            escenario.get('monto_8', ''),
            escenario.get('monto_9', ''),
            escenario.get('monto_10', ''),
            escenario.get('monto_11', ''),
            escenario.get('monto_12', ''),
            escenario.get('factor_8', ''),
            escenario.get('factor_9', ''),
            escenario.get('factor_10', ''),
            escenario.get('factor_11', ''),
            escenario.get('factor_12', ''),
            escenario['numero_dj'],
            escenario['fecha_informe'],
            escenario['observaciones'],
        ]
        ws.append(fila)
    
    # Aplicar formato
    aplicar_formato_excel(ws)
    
    # Guardar archivo
    wb.save(OUTPUT_PATH)
    
    print(f"\n{'='*70}")
    print(f"✓ Archivo generado exitosamente: {OUTPUT_PATH}")
    print(f"{'='*70}")


# ==============================================================================
# PREPARACIÓN DEL ENTORNO
# ==============================================================================

def preparar_instrumentos():
    """
    Crea los instrumentos financieros necesarios para las pruebas.
    """
    print(f"\n{'='*70}")
    print(f"PREPARANDO INSTRUMENTOS DE PRUEBA")
    print(f"{'='*70}")
    
    instrumentos = [
        ('TEST001', 'Instrumento de Prueba 001', 'Bono'),
        ('TEST002', 'Instrumento de Prueba 002', 'Acción'),
    ]
    
    for codigo, nombre, tipo in instrumentos:
        crear_instrumento_prueba(codigo, nombre, tipo)


# ==============================================================================
# FUNCIÓN PRINCIPAL
# ==============================================================================

def main():
    """
    Función principal del script.
    """
    print(f"\n{'#'*70}")
    print(f"#  SCRIPT DE PRUEBA FINAL - VALIDACIONES Y DUPLICADOS")
    print(f"#  Fecha: {date.today()}")
    print(f"{'#'*70}")
    
    try:
        # Paso 1: Preparar instrumentos
        preparar_instrumentos()
        
        # Paso 2: Generar archivo Excel
        generar_excel_prueba()
        
        # Resumen final
        print(f"\n{'='*70}")
        print(f"RESUMEN DE PRUEBAS")
        print(f"{'='*70}")
        print(f"✓ Archivo creado: {OUTPUT_FILE}")
        print(f"✓ Total de registros: 3")
        print(f"")
        print(f"RESULTADOS ESPERADOS AL CARGAR EL ARCHIVO:")
        print(f"  - Fila 2 (GOLDEN):     ✓ Éxito (1 registro)")
        print(f"  - Fila 3 (MATH FAIL):  ✗ Error de validación")
        print(f"  - Fila 4 (DUPLICATE):  ✗ Error de duplicado")
        print(f"")
        print(f"ESTADÍSTICAS ESPERADAS:")
        print(f"  - Procesados: 3")
        print(f"  - Exitosos: 1")
        print(f"  - Fallidos: 2")
        print(f"  - Estado: PARCIAL")
        print(f"")
        print(f"PRÓXIMOS PASOS:")
        print(f"  1. Ir a http://127.0.0.1:8000/carga-masiva/")
        print(f"  2. Subir el archivo: {OUTPUT_FILE}")
        print(f"  3. Verificar que solo 1 registro se carga exitosamente")
        print(f"  4. Revisar errores detallados en el historial")
        print(f"{'='*70}\n")
        
        return 0
        
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit_code = main()
    sys.exit(exit_code)
