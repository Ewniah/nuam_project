"""
Script para visualizar el contenido del archivo Excel generado.
"""

import openpyxl
from datetime import datetime


def mostrar_contenido_excel():
    """
    Muestra el contenido del archivo Excel de manera legible.
    """
    archivo = "datos_prueba_carga_masiva.xlsx"
    
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        
        print("=" * 100)
        print(f"📄 CONTENIDO DE: {archivo}")
        print("=" * 100)
        print()
        
        # Headers
        headers = [cell.value for cell in sheet[1]]
        print("📋 COLUMNAS:")
        for i, header in enumerate(headers, 1):
            print(f"   {i}. {header}")
        print()
        
        # Datos
        print("📊 DATOS (Primeros 10 registros):")
        print("-" * 100)
        
        registros = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), 1):
            if row[0]:  # Si tiene código de instrumento
                registros.append(row)
                print(f"\nRegistro #{i}:")
                print(f"   Código Instrumento: {row[0]}")
                print(f"   Nombre: {row[1]}")
                print(f"   Tipo: {row[2]}")
                print(f"   Monto: ${row[3]:,.2f}" if row[3] else f"   Monto: -")
                print(f"   Factor: {row[4]}" if row[4] else f"   Factor: -")
                print(f"   Método: {row[5]}")
                print(f"   DJ: {row[6]}")
                print(f"   Fecha Informe: {row[7]}")
                if row[8]:
                    print(f"   Observaciones: {row[8]}")
        
        print()
        print("-" * 100)
        print(f"✅ Total de registros en archivo: {sum(1 for row in sheet.iter_rows(min_row=2) if row[0].value)}")
        print()
        
        # Resumen por tipo
        print("📈 RESUMEN POR TIPO DE INSTRUMENTO:")
        tipos = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                tipo = row[2] or "Sin tipo"
                tipos[tipo] = tipos.get(tipo, 0) + 1
        
        for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
            print(f"   • {tipo}: {count} registros")
        
        print()
        
        # Resumen por método
        print("📈 RESUMEN POR MÉTODO DE INGRESO:")
        metodos = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                metodo = row[5] or "Sin método"
                metodos[metodo] = metodos.get(metodo, 0) + 1
        
        for metodo, count in sorted(metodos.items(), key=lambda x: x[1], reverse=True):
            print(f"   • {metodo}: {count} registros")
        
        print()
        print("=" * 100)
        print("🎯 ARCHIVO LISTO PARA CARGA MASIVA")
        print("=" * 100)
        print()
        print("📍 Siguiente paso: Ir a http://127.0.0.1:8000/carga-masiva/")
        print("📤 Subir el archivo: datos_prueba_carga_masiva.xlsx")
        print()
        
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo: {archivo}")
        print("   Ejecuta primero: python scripts/generar_datos_prueba.py")
    except Exception as e:
        print(f"❌ Error al leer el archivo: {e}")


if __name__ == "__main__":
    mostrar_contenido_excel()
